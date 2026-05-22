"""
Dashboard del cliente: marcas registradas, historial de consultas, vigilancia.

Todas las rutas requieren login. La página /dashboard renderiza un SPA simple
con tabs (Alpine.js + fetch) y los datos vienen por API JSON.
"""

from __future__ import annotations

import logging
import os
from datetime import datetime

from flask import Blueprint, jsonify, render_template_string, request

from database import (
    AlertaVigilancia, Consulta, MarcaCliente, Pago,
    SuscripcionVigilancia, UserProfile, get_session,
)
from services.auth import current_user, has_active_premium, login_required, premium_required

logger = logging.getLogger(__name__)
bp = Blueprint("dashboard", __name__)

PRECIO_VIGILANCIA_MARCA = float(os.getenv("PRECIO_VIGILANCIA_INDIVIDUAL",
                                           os.getenv("PRECIO_VIGILANCIA_MARCA", "4900")))
PRECIO_VIGILANCIA_MULTI = float(os.getenv("PRECIO_VIGILANCIA_MULTI", "9900"))
PRECIO_VIGILANCIA_PORTFOLIO = float(os.getenv("PRECIO_VIGILANCIA_PORTFOLIO", "20000"))


def _ok(data, status=200):
    return jsonify({"ok": True, "data": data}), status


def _err(msg, status=400):
    return jsonify({"ok": False, "error": msg}), status


# ─────────────────────────────────────────────────────────────────────
# Página HTML del dashboard (Alpine.js)
# ─────────────────────────────────────────────────────────────────────

DASHBOARD_PAGE = """<!DOCTYPE html>
<html lang="es"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Mi panel — LegalPacers</title>
<script defer src="https://unpkg.com/alpinejs@3.13.0/dist/cdn.min.js"></script>
<style>
  *{box-sizing:border-box}
  body{margin:0;font-family:system-ui,-apple-system,sans-serif;background:#F4F5F9;color:#0D1B4B}
  .nav{background:#fff;border-bottom:1px solid #E2E8F0;padding:14px 32px;
       display:flex;align-items:center;justify-content:space-between}
  .logo{font-weight:800;font-size:18px;letter-spacing:.5px}
  .logo .accent{color:#1B6EF3}
  .user{font-size:14px;color:#64748b}
  .user a{color:#DC2626;margin-left:12px;text-decoration:none}
  main{max-width:1100px;margin:32px auto;padding:0 24px}
  h1{margin:0 0 24px}
  .tabs{display:flex;gap:0;background:#fff;border-radius:10px;padding:4px;margin-bottom:24px;
        box-shadow:0 1px 3px rgba(0,0,0,.05);max-width:fit-content}
  .tab{padding:10px 20px;border-radius:6px;cursor:pointer;font-weight:600;font-size:14px;
       color:#64748b;transition:.15s}
  .tab.active{background:#1B6EF3;color:#fff}
  .card{background:#fff;border-radius:12px;padding:24px;margin-bottom:16px;
        box-shadow:0 1px 4px rgba(0,0,0,.05)}
  .badge{display:inline-block;padding:2px 10px;border-radius:99px;font-size:12px;font-weight:600}
  .badge.green{background:#DCFCE7;color:#16A34A}
  .badge.yellow{background:#FEF9C3;color:#B45309}
  .badge.red{background:#FEE2E2;color:#DC2626}
  .badge.gray{background:#E5E7EB;color:#374151}
  .grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(280px,1fr));gap:16px}
  table{width:100%;border-collapse:collapse;font-size:14px}
  table td,table th{padding:10px;text-align:left;border-bottom:1px solid #E2E8F0}
  table th{font-size:12px;color:#64748b;text-transform:uppercase;font-weight:600}
  button{padding:10px 18px;background:#1B6EF3;color:#fff;border:none;border-radius:8px;
         font-weight:600;cursor:pointer;font-size:14px}
  button:hover{background:#1656c4}
  button.sec{background:#fff;color:#1B6EF3;border:1px solid #1B6EF3}
  button.danger{background:#DC2626}
  button.small{padding:6px 12px;font-size:13px}
  input,select{padding:10px;border:1px solid #E2E8F0;border-radius:8px;font-size:14px;
               width:100%;font-family:inherit}
  label{display:block;font-size:13px;font-weight:600;margin:10px 0 4px;color:#0D1B4B}
  .empty{text-align:center;padding:40px;color:#64748b}
  .alert-row{padding:14px;border-radius:8px;margin-bottom:8px;display:flex;justify-content:space-between;align-items:center}
  .alert-row.alto{background:#FEE2E2}
  .alert-row.medio{background:#FEF9C3}
  .alert-row.bajo{background:#DBEAFE}
  .modal{position:fixed;inset:0;background:rgba(13,27,75,.4);display:flex;align-items:center;
         justify-content:center;z-index:50;padding:20px}
  .modal-content{background:#fff;border-radius:12px;padding:28px;max-width:480px;width:100%}
</style></head>
<body x-data="dashboard()" x-init="cargar()">

<div class="nav">
  <div class="logo">LEGAL<span class="accent">PACERS</span> · Mi panel</div>
  <div class="user" style="position:relative" @click.outside="userMenuOpen=false">
    <div @click="userMenuOpen=!userMenuOpen" role="button" tabindex="0"
         style="display:inline-flex;align-items:center;gap:10px;background:#fff;border:1px solid #E2E8F0;
                cursor:pointer;padding:8px 14px;border-radius:10px;color:#0D1B4B;font-family:inherit;user-select:none"
         :style="userMenuOpen ? 'background:#F4F5F9;border-color:#CBD5E1' : 'background:#fff'">
      <div style="text-align:right;line-height:1.2">
        <div style="font-size:13px;font-weight:600;color:#0D1B4B"
             x-text="user.nombre || user.email.split('@')[0]"></div>
        <div style="font-size:11px;color:#64748b" x-text="user.email"></div>
      </div>
      <span style="color:#64748b;font-size:10px;margin-left:2px">▼</span>
    </div>

    <div x-show="userMenuOpen" x-cloak
         style="position:absolute;right:0;top:calc(100% + 6px);background:#fff;border:1px solid #E2E8F0;
                border-radius:10px;box-shadow:0 6px 20px rgba(13,27,75,.12);min-width:240px;
                padding:6px;z-index:30">
      <div style="padding:10px 12px;border-bottom:1px solid #E2E8F0">
        <div style="font-size:13px;font-weight:600;color:#0D1B4B"
             x-text="user.nombre || user.email.split('@')[0]"></div>
        <div style="font-size:12px;color:#64748b" x-text="user.email"></div>
        <span class="badge" :class="user.is_admin ? 'green' : 'gray'" style="margin-top:6px;font-size:10px"
              x-text="user.is_admin ? 'Admin' : 'Premium'"></span>
      </div>
      <a @click.prevent="tab='perfil'; userMenuOpen=false" href="#"
         style="display:block;padding:10px 12px;color:#0D1B4B;text-decoration:none;font-size:14px;border-radius:6px"
         onmouseover="this.style.background='#F4F5F9'" onmouseout="this.style.background='transparent'">
        Mi perfil
      </a>
      <a @click.prevent="modalPassword=true; userMenuOpen=false" href="#"
         style="display:block;padding:10px 12px;color:#0D1B4B;text-decoration:none;font-size:14px;border-radius:6px"
         onmouseover="this.style.background='#F4F5F9'" onmouseout="this.style.background='transparent'">
        Cambiar contraseña
      </a>
      <hr style="border:none;border-top:1px solid #E2E8F0;margin:4px 0">
      <a href="/logout"
         style="display:block;padding:10px 12px;color:#DC2626;text-decoration:none;font-size:14px;border-radius:6px"
         onmouseover="this.style.background='#FEE2E2'" onmouseout="this.style.background='transparent'">
        Cerrar sesión
      </a>
    </div>
  </div>
</div>

<main>
  <h1>Hola<span x-show="user.nombre" x-text="', ' + user.nombre"></span> 👋</h1>

  <div class="tabs">
    <div class="tab" :class="tab==='buscar'&&'active'" @click="tab='buscar'">Buscar marca</div>
    <div class="tab" :class="tab==='consultas'&&'active'" @click="tab='consultas'">Consultas</div>
    <div class="tab" :class="tab==='marcas'&&'active'" @click="tab='marcas'">Mis marcas</div>
    <div class="tab" :class="tab==='vigilancia'&&'active'" @click="tab='vigilancia'">Vigilancia</div>
    <div class="tab" :class="tab==='alertas'&&'active'" @click="tab='alertas'">Alertas</div>
    <div class="tab" :class="tab==='pagos'&&'active'" @click="tab='pagos'">Pagos</div>
    <div class="tab" :class="tab==='perfil'&&'active'" @click="tab='perfil'">Mi perfil</div>
    <div class="tab" :class="tab==='config'&&'active'" @click="tab='config'">Configuración</div>
    <div x-show="user.is_admin" class="tab" :class="tab==='notorias'&&'active'"
         @click="tab='notorias'; fetchNotorias()" style="background:#FEF3C7;color:#92400E">
      ⭐ Notorias <span style="font-size:11px">(admin)</span>
    </div>
  </div>

  <!-- BUSCAR (premium) -->
  <div x-show="tab==='buscar'" class="card">
    <h3 style="margin-top:0">Búsqueda Premium</h3>
    <p style="color:#64748b;margin:0 0 20px">
      Consultas y análisis ilimitados. Te mostramos las coincidencias completas con
      titular, clase y nivel de confundibilidad (gráfica, fonética, ideológica).
    </p>

    <div style="display:grid;grid-template-columns:2fr 1fr;gap:12px">
      <div>
        <label>Marca a consultar *</label>
        <input type="text" name="marca-search" x-model="buscar.marca"
               placeholder="Ej: Acme, MiMarca..." autocomplete="off"
               spellcheck="false" autocorrect="off" autocapitalize="off"
               @keydown.enter="ejecutarBusqueda()">
      </div>
      <div>
        <label>Clases Niza
          <span style="font-weight:400;color:#64748b;font-size:11px">
            (tildá una o varias; sin tildar = todas)
          </span>
        </label>
        <div style="border:1px solid #E2E8F0;border-radius:8px;padding:8px;max-height:180px;overflow-y:auto;background:#fff">
          <div style="display:flex;gap:6px;margin-bottom:6px">
            <button class="small sec" type="button" @click.prevent="buscar.clases=[]"
                    style="padding:4px 10px;font-size:12px">Ninguna</button>
            <button class="small sec" type="button" @click.prevent="buscar.clases=Array.from({length:45},(_,i)=>String(i+1))"
                    style="padding:4px 10px;font-size:12px">Todas</button>
            <input type="text" x-model="buscar.claseFiltro" placeholder="Filtrar..."
                   style="flex:1;padding:4px 8px;font-size:12px" autocomplete="off">
          </div>
          <template x-for="n in 45" :key="n">
            <label x-show="!buscar.claseFiltro || ('' + n + ' ' + (NIZA_TITLES[n]||'')).toLowerCase().includes(buscar.claseFiltro.toLowerCase())"
                   style="display:flex;align-items:center;gap:8px;padding:4px 6px;cursor:pointer;font-weight:400;margin:0;border-radius:4px"
                   onmouseover="this.style.background='#F4F5F9'" onmouseout="this.style.background='transparent'">
              <input type="checkbox" :value="String(n)" x-model="buscar.clases" style="width:auto;margin:0">
              <span style="font-size:13px">
                <strong x-text="n"></strong> · <span x-text="NIZA_TITLES[n] || ''"></span>
              </span>
            </label>
          </template>
        </div>
        <p style="font-size:11px;color:#64748b;margin:4px 0 0">
          <span x-show="buscar.clases.length === 0">Buscaremos en las 45 clases.</span>
          <span x-show="buscar.clases.length > 0">
            <strong x-text="buscar.clases.length"></strong> clase(s) seleccionada(s).
          </span>
        </p>
      </div>
    </div>

    <label style="margin-top:12px">Descripción del producto / servicio (opcional)</label>
    <input type="text" name="marca-descripcion" x-model="buscar.descripcion"
           placeholder="Ayuda al análisis conceptual. Ej: línea de gaseosas"
           autocomplete="off" spellcheck="false">

    <button @click="ejecutarBusqueda()" :disabled="buscando" style="margin-top:18px">
      <span x-show="!buscando">Buscar →</span>
      <span x-show="buscando" x-cloak>Buscando…</span>
    </button>

    <div x-show="buscarErr" class="alert-row alto" style="margin-top:14px" x-text="buscarErr"></div>

    <template x-if="buscarResult">
      <div style="margin-top:24px">
        <!-- VEREDICTO PRINCIPAL: marca notoria primero, después el resto -->
        <div x-show="buscarResult.es_notoria"
             style="background:#FEE2E2;border:2px solid #DC2626;border-radius:12px;
                    padding:20px;margin-bottom:18px">
          <div style="font-size:22px;font-weight:800;color:#991B1B;margin-bottom:8px">
            🛑 MARCA NOTORIA — No registrable
          </div>
          <p style="margin:0 0 8px;line-height:1.5;color:#7F1D1D" x-html="buscarResult.mensaje"></p>
          <p style="margin:0;font-size:14px;color:#7F1D1D">
            Aunque el análisis técnico completo igual lo hacemos abajo, te adelantamos:
            <strong>las marcas notorias tienen protección en las 45 clases</strong>. Cualquier
            registro confusable va a ser rechazado, sin importar el rubro.
          </p>
        </div>

        <div x-show="!buscarResult.es_notoria" class="alert-row" :class="{
          alto: buscarResult.veredicto==='no_disponible',
          medio: buscarResult.veredicto==='necesita_analisis',
          bajo: buscarResult.veredicto==='probablemente_disponible'
        }" style="display:block">
          <div style="font-size:18px;font-weight:700;margin-bottom:4px"
               x-text="{
                 probablemente_disponible:'✓ Buenas señales',
                 necesita_analisis:'⚠ Necesita análisis',
                 no_disponible:'✕ Riesgo alto'
               }[buscarResult.veredicto]"></div>
          <div x-html="buscarResult.mensaje"></div>
        </div>

        <!-- INFORME RESUMIDO -->
        <div style="margin-top:18px;padding:18px;background:#fff;border:1px solid #E2E8F0;border-radius:10px">
          <h4 style="margin:0 0 12px">Informe de búsqueda</h4>

          <p style="margin:0 0 12px;line-height:1.6">
            Para <strong x-text="'&quot;' + buscarResult.marca + '&quot;'"></strong>
            <span x-show="buscarResult.clases_consultadas?.length">
              en
              <span x-show="buscarResult.clases_consultadas.length === 1">
                la clase <strong x-text="buscarResult.clases_consultadas[0]"></strong>
              </span>
              <span x-show="buscarResult.clases_consultadas.length > 1">
                las clases <strong x-text="buscarResult.clases_consultadas.join(', ')"></strong>
              </span>:
            </span>
            <span x-show="!buscarResult.clases_consultadas?.length">en las 45 clases:</span>
          </p>

          <ul style="margin:0;padding-left:22px;line-height:1.8">
            <li>
              Encontramos <strong x-text="buscarResult.stats.matches_total"></strong>
              marca(s) similar(es) ya registrada(s) en Argentina
              <span x-show="buscarResult.stats.identicas > 0">
                — <strong style="color:#DC2626" x-text="buscarResult.stats.identicas"></strong> idéntica(s) o casi idéntica(s)
              </span>.
            </li>
            <li>
              <span style="color:#1B6EF3">●</span>
              <strong>Léxicamente</strong> (cómo se escribe):
              <strong x-text="buscarResult.stats.similares_lex"></strong> con similitud ≥ 50%,
              de las cuales <strong x-text="buscarResult.stats.similares_lex_alto || 0"></strong> son muy similares (≥ 70%).
            </li>
            <li>
              <span style="color:#7C3AED">●</span>
              <strong>Fonéticamente</strong> (cómo suena):
              <strong x-text="buscarResult.stats.similares_fon"></strong> con similitud ≥ 50%,
              <strong x-text="buscarResult.stats.similares_fon_alto || 0"></strong> ≥ 70%.
            </li>
            <li>
              <span style="color:#16A34A">●</span>
              <strong>Conceptualmente</strong> (significado, sinónimos, traducciones):
              <strong x-text="buscarResult.stats.similares_con"></strong> con similitud ≥ 50%,
              <strong x-text="buscarResult.stats.similares_con_alto || 0"></strong> ≥ 70%.
            </li>
            <li>
              <strong>¿Es marca notoria?</strong>
              <span x-show="buscarResult.es_notoria" style="color:#DC2626;font-weight:700">
                Sí — atención, hay alto riesgo (ver alerta roja abajo)
              </span>
              <span x-show="!buscarResult.es_notoria" style="color:#16A34A">No detectada.</span>
            </li>
          </ul>
          <p style="font-size:12px;color:#64748b;margin:12px 0 0">
            Nota: una marca puede tener nivel "alto" en el score final aunque ninguna
            dimensión llegue al 70% individual, gracias a los bonus +10% si coincide
            la clase y +5% si es vigente.
          </p>
        </div>

        <!-- MARCA FUERTE / DÉBIL -->
        <div x-show="buscarResult.marca_strength" style="margin-top:18px;padding:18px;
                    background:#fff;border:1px solid #E2E8F0;border-radius:10px">
          <h4 style="margin:0 0 6px">¿Tu marca es fuerte o débil?</h4>
          <p style="margin:0 0 10px;font-size:13px;color:#64748b">
            Una marca <strong>fuerte</strong> es distintiva (inventada o sin elementos comunes del rubro)
            y tiene amplia protección legal. Una marca <strong>débil</strong> contiene prefijos / sufijos
            de uso común (Rapi-, Eco-, -farma) o palabras genéricas, y tiene protección acotada.
          </p>
          <div style="display:flex;gap:14px;align-items:center;margin:12px 0">
            <div :style="(buscarResult.marca_strength?.clasificacion === 'fuerte' ? 'background:#DCFCE7;color:#16A34A' : (buscarResult.marca_strength?.clasificacion === 'debil' ? 'background:#FEE2E2;color:#DC2626' : 'background:#FEF3C7;color:#92400E')) + ';padding:14px 22px;border-radius:10px;font-weight:700;font-size:16px;text-transform:uppercase'"
                 x-text="buscarResult.marca_strength?.clasificacion"></div>
            <div style="flex:1">
              <div style="font-size:14px;color:#64748b">Puntaje de distintividad</div>
              <div style="height:10px;background:#E2E8F0;border-radius:99px;overflow:hidden;margin-top:4px">
                <div :style="'width:' + (buscarResult.marca_strength?.puntaje || 0) + '%; height:100%; background:' + (buscarResult.marca_strength?.puntaje >= 75 ? '#16A34A' : (buscarResult.marca_strength?.puntaje >= 45 ? '#D97706' : '#DC2626'))"></div>
              </div>
              <div style="font-size:13px;font-weight:600;margin-top:4px"
                   x-text="(buscarResult.marca_strength?.puntaje || 0) + ' / 100'"></div>
            </div>
          </div>
          <ul style="margin:8px 0 0;padding-left:22px;line-height:1.7;font-size:13px;color:#475569">
            <template x-for="r in (buscarResult.marca_strength?.razones || [])" :key="r">
              <li x-text="r"></li>
            </template>
          </ul>
          <p x-show="buscarResult.mot_vedette" style="font-size:13px;color:#64748b;margin-top:12px;
                    padding:8px 12px;background:#F4F5F9;border-radius:6px">
            <strong>Elemento predominante (Mot Vedette):</strong>
            <span x-text="buscarResult.mot_vedette" style="font-weight:600;color:#1B6EF3"></span>
            — Es la palabra que más capta atención. El INPI lo usa para comparar marcas con varios términos.
          </p>
        </div>

        <!-- MARCAS VENCIDAS -->
        <div x-show="buscarResult.marcas_vencidas && buscarResult.marcas_vencidas.length"
             style="margin-top:18px;padding:18px;background:#fff;border:1px solid #E2E8F0;border-radius:10px">
          <h4 style="margin:0 0 6px">Marcas vencidas o no vigentes</h4>
          <p style="margin:0 0 10px;font-size:13px;color:#64748b">
            Estas marcas similares <strong>ya no están activas</strong> (vencieron o cayeron en abandono).
            Volvieron al dominio público y podrían reusarse — aunque siempre conviene confirmar el estado actual.
          </p>
          <div style="overflow-x:auto">
          <table>
            <thead><tr>
              <th>Marca</th><th>Clase</th><th>Titular original</th><th>Estado</th><th>Vencimiento</th>
            </tr></thead>
            <tbody>
              <template x-for="v in buscarResult.marcas_vencidas" :key="'venc-'+v.denominacion+v.clase">
                <tr>
                  <td><strong x-text="v.denominacion"></strong></td>
                  <td x-text="v.clase || '—'"></td>
                  <td x-text="v.titular || '—'"></td>
                  <td><span class="badge gray" x-text="v.estado"></span></td>
                  <td x-text="fmtDate(v.fecha_vencimiento) || '—'"></td>
                </tr>
              </template>
            </tbody>
          </table>
          </div>
        </div>

        <!-- ANÁLISIS TEMPORAL -->
        <div x-show="buscarResult.temporal && buscarResult.temporal.por_anio && buscarResult.temporal.por_anio.length"
             style="margin-top:18px;padding:18px;background:#fff;border:1px solid #E2E8F0;border-radius:10px">
          <h4 style="margin:0 0 6px">Actividad reciente en tu clase</h4>
          <p style="margin:0 0 10px;font-size:13px;color:#64748b">
            Cuántas marcas se registraron por año en las clases consultadas — para saber si tu rubro
            está saturado o despejado de marcas nuevas.
          </p>
          <div style="display:flex;align-items:end;gap:6px;height:120px;padding:10px 0">
            <template x-for="t in (buscarResult.temporal?.por_anio || [])" :key="'tmp-'+t.anio">
              <div style="flex:1;display:flex;flex-direction:column;align-items:center;gap:4px">
                <div style="font-size:11px;color:#64748b" x-text="t.registros"></div>
                <div :style="'width:80%;background:#1B6EF3;height:' + Math.max(4, Math.min(100, t.registros * 100 / (buscarResult.temporal?.promedio_anual * 2 || 100))) + 'px;border-radius:4px 4px 0 0'"></div>
                <div style="font-size:11px;font-weight:600" x-text="t.anio"></div>
              </div>
            </template>
          </div>
          <div style="font-size:13px;color:#64748b;text-align:center;margin-top:6px">
            Promedio: <strong x-text="buscarResult.temporal?.promedio_anual + ' marcas/año'"></strong>
            · Total últimos 5 años:
            <strong x-text="buscarResult.temporal?.total_5_anios"></strong>
          </div>
        </div>

        <!-- CLASES SUGERIDAS -->
        <div x-show="buscarResult.clases_sugeridas && buscarResult.clases_sugeridas.length"
             style="margin-top:18px;padding:18px;background:#fff;border:1px solid #E2E8F0;border-radius:10px">
          <h4 style="margin:0 0 6px">Clases adicionales que conviene considerar</h4>
          <p style="margin:0 0 10px;font-size:13px;color:#64748b">
            Según tu descripción, estas clases también podrían serte útiles. Registrar en varias
            clases protege mejor tu marca contra usos no autorizados.
          </p>
          <div style="display:flex;flex-direction:column;gap:8px">
            <template x-for="c in buscarResult.clases_sugeridas" :key="'sug-'+c.clase">
              <div style="padding:10px 14px;background:#F0F5FF;border-radius:8px;
                          display:flex;justify-content:space-between;align-items:center;gap:14px">
                <div style="flex:1">
                  <strong>Clase <span x-text="c.clase"></span></strong>
                  <span style="color:#64748b;font-size:13px"
                        x-text="' — ' + (NIZA_TITLES[c.clase] || '')"></span>
                  <div style="font-size:12px;color:#475569;margin-top:2px"
                       x-text="c.razon"></div>
                </div>
              </div>
            </template>
          </div>
        </div>

        <!-- PROBABILIDAD POR CLASE -->
        <div x-show="buscarResult.por_clase" style="margin-top:18px;padding:18px;
                    background:#fff;border:1px solid #E2E8F0;border-radius:10px">
          <h4 style="margin:0 0 6px">Probabilidad de registro por clase</h4>
          <p style="margin:0 0 14px;font-size:13px;color:#64748b">
            Ponderamos las coincidencias detectadas por clase para estimar la probabilidad
            de registro: matches altos restan 25 pts, medios 10 pts y bajos 3 pts.
          </p>

          <!-- Si el usuario eligió clases puntuales, mostramos sólo esas como foco -->
          <div x-show="buscarResult.clases_consultadas && buscarResult.clases_consultadas.length > 0
                       && buscarResult.clases_consultadas.length < 45"
               style="margin-bottom:14px">
            <div style="font-weight:600;margin-bottom:6px">Tus clases seleccionadas</div>
            <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(240px,1fr));gap:6px">
              <template x-for="s in (buscarResult.por_clase?.scores || []).filter(x => x.es_pedida)" :key="'sel-'+s.clase">
                <div style="background:#DBEAFE;padding:8px 12px;border-radius:8px">
                  <strong style="font-size:13px" x-text="s.clase + '. ' + s.titulo"></strong>
                  <div style="display:flex;justify-content:space-between;margin-top:4px;font-size:12px">
                    <span style="color:#64748b" x-text="s.matches + ' coincidencias'"></span>
                    <strong :style="s.label === 'alta' ? 'color:#16A34A' : (s.label === 'media' ? 'color:#D97706' : 'color:#DC2626')"
                            x-text="s.probabilidad + '% prob.'"></strong>
                  </div>
                </div>
              </template>
            </div>
          </div>

          <!-- Si el usuario seleccionó todas (o ninguna = todas), mostrar mejores/peores -->
          <div x-show="!buscarResult.clases_consultadas?.length
                        || buscarResult.clases_consultadas?.length === 45"
               style="display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:14px">
            <div>
              <div style="font-weight:600;color:#16A34A;margin-bottom:6px">Mejores clases para registrar</div>
              <template x-for="m in (buscarResult.por_clase?.mejores || [])" :key="'best-'+m.clase">
                <div style="display:flex;justify-content:space-between;padding:4px 0;font-size:13px">
                  <span><strong x-text="m.clase"></strong> · <span x-text="m.titulo"></span></span>
                  <strong style="color:#16A34A" x-text="m.probabilidad + '%'"></strong>
                </div>
              </template>
            </div>
            <div>
              <div style="font-weight:600;color:#DC2626;margin-bottom:6px">Clases con mayor riesgo</div>
              <template x-for="p in (buscarResult.por_clase?.peores || [])" :key="'wor-'+p.clase">
                <div style="display:flex;justify-content:space-between;padding:4px 0;font-size:13px">
                  <span><strong x-text="p.clase"></strong> · <span x-text="p.titulo"></span></span>
                  <strong style="color:#DC2626" x-text="p.probabilidad + '%'"></strong>
                </div>
              </template>
            </div>
          </div>

          <details>
            <summary style="cursor:pointer;font-size:13px;color:#1B6EF3;font-weight:600">
              Ver todas las 45 clases
            </summary>
            <div style="margin-top:10px;display:grid;grid-template-columns:repeat(auto-fill,minmax(220px,1fr));gap:6px">
              <template x-for="s in (buscarResult.por_clase?.scores || [])" :key="'all-'+s.clase">
                <div :style="(s.es_pedida ? 'background:#DBEAFE;' : 'background:#F8FAFC;') + 'padding:6px 10px;border-radius:6px;font-size:12px'">
                  <strong x-text="s.clase + '. ' + s.titulo"></strong>
                  <div style="display:flex;justify-content:space-between;margin-top:2px">
                    <span style="color:#64748b" x-text="s.matches + ' matches'"></span>
                    <strong :style="s.label === 'alta' ? 'color:#16A34A' : (s.label === 'media' ? 'color:#D97706' : 'color:#DC2626')"
                            x-text="s.probabilidad + '%'"></strong>
                  </div>
                </div>
              </template>
            </div>
          </details>
        </div>

        <div x-show="buscarResult.dominios && buscarResult.dominios.length" style="margin-top:16px">
          <div style="font-weight:600;margin-bottom:6px">Dominios web</div>
          <template x-for="d in buscarResult.dominios" :key="d.domain">
            <span class="badge" :class="d.status==='disponible'?'green':(d.status==='tomado'?'red':'gray')"
                  style="margin-right:6px;margin-bottom:4px;display:inline-block">
              <span x-text="d.domain"></span> · <span x-text="d.status"></span>
            </span>
          </template>
        </div>

        <div x-show="buscarResult.handles && buscarResult.handles.length" style="margin-top:14px">
          <div style="font-weight:600;margin-bottom:6px">Usuarios en redes sociales</div>
          <template x-for="h in buscarResult.handles" :key="h.plataforma + h.handle">
            <a :href="h.url" target="_blank" rel="noopener"
               class="badge" :class="h.status==='disponible'?'green':(h.status==='tomado'?'red':'gray')"
               style="margin-right:6px;margin-bottom:4px;display:inline-block;text-decoration:none">
              <strong x-text="h.plataforma"></strong>
              <span x-text="h.handle"></span> ·
              <span x-text="h.status"></span>
            </a>
          </template>
          <p style="font-size:11px;color:#64748b;margin:6px 0 0">
            Chequeo informativo. Click para verificar manualmente.
          </p>
        </div>

        <div x-show="buscarResult.notorious_warnings && buscarResult.notorious_warnings.length"
             style="margin-top:20px;background:#FEE2E2;border:1px solid #FCA5A5;border-radius:10px;
                    padding:14px 16px">
          <div style="display:flex;align-items:center;gap:8px;font-weight:700;color:#991B1B">
            <span style="font-size:18px">🛑</span>
            Atención: marca notoria detectada
          </div>
          <p style="margin:6px 0 12px;font-size:13px;color:#7F1D1D">
            Tu búsqueda se parece a una marca notoria. Las marcas notorias tienen
            <strong>protección extendida a todas las clases</strong>: no podés registrarla
            ni siquiera en otro rubro porque sería aprovechamiento parasitario.
          </p>
          <div style="overflow-x:auto">
          <table>
            <thead><tr>
              <th>Marca notoria</th><th>Score lex</th><th>Score fon</th><th>Score final</th>
            </tr></thead>
            <tbody>
              <template x-for="n in buscarResult.notorious_warnings" :key="'not-'+n.denominacion">
                <tr>
                  <td><strong x-text="n.denominacion"></strong></td>
                  <td x-text="((n.scores?.lexical||0)*100).toFixed(0)+'%'"></td>
                  <td x-text="((n.scores?.fonetica||0)*100).toFixed(0)+'%'"></td>
                  <td><strong x-text="((n.score||0)*100).toFixed(0)+'%'"></strong></td>
                </tr>
              </template>
            </tbody>
          </table>
          </div>
        </div>

        <div x-show="buscarResult.cross_class_matches && buscarResult.cross_class_matches.length"
             style="margin-top:20px;background:#FEF3C7;border:1px solid #FDE68A;border-radius:10px;
                    padding:14px 16px">
          <div style="display:flex;align-items:center;gap:8px;font-weight:700;color:#92400E">
            <span style="font-size:18px">⚠️</span>
            Marcas similares en otras clases — atención si son notorias
          </div>
          <p style="margin:6px 0 12px;font-size:13px;color:#78350F">
            Aunque no aparezcan en tu clase, la ley argentina de marcas protege las marcas notorias
            (Coca-Cola, Nike, etc.) <strong>también en clases distintas</strong>. Si alguna de estas
            es notoria, no podrías registrar una marca confusable, incluso en otra clase.
          </p>
          <div style="overflow-x:auto">
          <table>
            <thead><tr>
              <th>Marca</th><th>Clase</th><th>Titular</th><th>Score</th>
            </tr></thead>
            <tbody>
              <template x-for="m in buscarResult.cross_class_matches" :key="'cc-'+m.id">
                <tr>
                  <td><strong x-text="m.denominacion"></strong></td>
                  <td>
                    <span x-text="m.clase"></span> ·
                    <span style="font-size:11px;color:#64748b"
                          x-text="NIZA_TITLES[m.clase] || ''"></span>
                  </td>
                  <td x-text="m.titular || '—'"></td>
                  <td><strong x-text="((m.score||0)*100).toFixed(0) + '%'"></strong></td>
                </tr>
              </template>
            </tbody>
          </table>
          </div>
        </div>

        <div x-show="buscarResult.matches && buscarResult.matches.length" style="margin-top:20px">
          <div style="display:flex;justify-content:space-between;align-items:start;gap:8px;margin:0 0 6px">
            <h4 style="margin:0">Marcas similares registradas en Argentina</h4>
            <button x-show="user.is_admin" class="small sec" @click="marcarTodasComoNotorias()"
                    style="font-size:11px;padding:4px 10px"
                    title="Admin: marcar las coincidencias 'alto' como notorias">
              ⭐ Marcar altos como notorios
            </button>
          </div>
          <p style="font-size:12px;color:#64748b;margin:0 0 12px">
            En Argentina se evalúa la <strong>confundibilidad</strong> entre marcas en
            3 dimensiones. Mostramos el score de cada una para que veas qué dispara el match.
          </p>
          <div style="overflow-x:auto">
          <table>
            <thead><tr>
              <th style="min-width:180px">Marca</th>
              <th>Clase</th>
              <th>Titular</th>
              <th>Estado</th>
              <th style="min-width:260px">Confundibilidad</th>
              <th>Score</th>
              <th>Nivel</th>
            </tr></thead>
            <tbody>
              <template x-for="m in buscarResult.matches" :key="m.id">
                <tr>
                  <td>
                    <strong x-text="m.denominacion"></strong>
                    <div style="font-size:11px;color:#64748b;margin-top:2px">
                      <span x-show="m.acta">Acta <span x-text="m.acta"></span></span>
                      <template x-for="tag in matchTags(m, buscar.marca, buscar.clase)" :key="tag">
                        <span class="badge yellow" style="margin-left:4px;font-size:10px"
                              x-text="tag"></span>
                      </template>
                    </div>
                  </td>
                  <td>
                    <span x-text="m.clase || '—'"></span>
                    <span class="badge green" style="margin-left:4px;font-size:10px"
                          x-show="buscar.clase && parseInt(buscar.clase) === m.clase">misma</span>
                  </td>
                  <td x-text="m.titular || '—'"></td>
                  <td><span class="badge gray" x-text="m.estado || m.estado_code || '—'"></span></td>
                  <td>
                    <div style="font-size:11px;line-height:1.4">
                      <div style="display:flex;align-items:center;gap:6px;margin-bottom:2px">
                        <span style="min-width:64px;color:#64748b">Léxico</span>
                        <div style="flex:1;background:#E2E8F0;border-radius:99px;height:6px;overflow:hidden">
                          <div :style="'width:'+((m.scores?.lexical||0)*100)+'%;height:100%;background:#1B6EF3'"></div>
                        </div>
                        <span style="min-width:34px;text-align:right;font-weight:600"
                              x-text="((m.scores?.lexical||0)*100).toFixed(0)+'%'"></span>
                      </div>
                      <div style="display:flex;align-items:center;gap:6px;margin-bottom:2px">
                        <span style="min-width:64px;color:#64748b">Fonético</span>
                        <div style="flex:1;background:#E2E8F0;border-radius:99px;height:6px;overflow:hidden">
                          <div :style="'width:'+((m.scores?.fonetica||0)*100)+'%;height:100%;background:#7C3AED'"></div>
                        </div>
                        <span style="min-width:34px;text-align:right;font-weight:600"
                              x-text="((m.scores?.fonetica||0)*100).toFixed(0)+'%'"></span>
                      </div>
                      <div style="display:flex;align-items:center;gap:6px">
                        <span style="min-width:64px;color:#64748b">Conceptual</span>
                        <div style="flex:1;background:#E2E8F0;border-radius:99px;height:6px;overflow:hidden">
                          <div :style="'width:'+((m.scores?.conceptual||0)*100)+'%;height:100%;background:#16A34A'"></div>
                        </div>
                        <span style="min-width:34px;text-align:right;font-weight:600"
                              x-text="((m.scores?.conceptual||0)*100).toFixed(0)+'%'"></span>
                      </div>
                      <div x-show="m.razon_conceptual" style="margin-top:4px;color:#475569;font-style:italic"
                           x-text="m.razon_conceptual"></div>
                    </div>
                  </td>
                  <td><strong x-text="((m.score||0)*100).toFixed(0) + '%'"></strong></td>
                  <td>
                    <span class="badge" :class="{
                      red: m.nivel==='alto', yellow: m.nivel==='medio', gray: m.nivel==='bajo'
                    }" x-text="m.nivel"></span>
                  </td>
                </tr>
              </template>
            </tbody>
          </table>
          </div>

          <div style="background:#F4F5F9;padding:12px 14px;border-radius:8px;margin-top:12px;
                      font-size:12px;color:#475569">
            <div style="display:flex;justify-content:space-between;align-items:center;cursor:pointer"
                 @click="ayudaScores=!ayudaScores">
              <strong>¿Cómo se calculan los porcentajes?</strong>
              <span x-text="ayudaScores ? '−' : '+'"
                    style="font-size:18px;color:#1B6EF3"></span>
            </div>

            <div x-show="ayudaScores" x-cloak style="margin-top:10px;line-height:1.6">
              <p style="margin:6px 0">
                <span style="color:#1B6EF3">●</span> <strong>Léxico (0-100%)</strong> —
                comparamos cómo se escriben las marcas: cuántas letras comparten y en qué orden.
                Una marca normalizada (sin tildes, sin caracteres especiales) se contrasta contra
                la otra; si son idénticas, 100%; si comparten núcleos largos, alto.
              </p>
              <p style="margin:6px 0">
                <span style="color:#7C3AED">●</span> <strong>Fonético (0-100%)</strong> —
                evaluamos cómo suenan las marcas en español. Aplicamos las reglas fonéticas locales
                (b↔v, c→k/s según vocal, h muda inicial, vocales internas, dobles letras) para
                identificar si dos marcas suenan igual aunque se escriban distinto.
                Ejemplo: "Hasúcar ≈ Azúcar".
              </p>
              <p style="margin:6px 0">
                <span style="color:#16A34A">●</span> <strong>Conceptual (0-100%)</strong> —
                a través de sistemas de IA evaluamos el <em>significado</em> de las marcas:
                detectamos sinónimos ("Los Criadores" ≈ "Los Ganaderos"), traducciones
                ("Norte" ≈ "Notte", "L'Etoile" ≈ "Stella"), antónimos que asocian ideas
                ("Fiel" ≈ "Infiel") y asociación de ideas en general.
              </p>
              <p style="margin:10px 0 6px;font-size:11px;color:#64748b">
                <strong>Score final</strong> = la mayor de las 3 dimensiones + bonus
                (+10% si comparten clase, +5% si la marca está "vigente" registrada). El
                <strong>nivel</strong> se asigna según el score combinado:
                rojo (alto) ≥ 75%, amarillo (medio) 60-75%, gris (bajo) 45-60%.
              </p>
            </div>
          </div>
        </div>

        <p x-show="(!buscarResult.matches || !buscarResult.matches.length)
                    && (!buscarResult.cross_class_matches || !buscarResult.cross_class_matches.length)
                    && (!buscarResult.notorious_warnings || !buscarResult.notorious_warnings.length)"
           style="margin-top:16px;color:#16A34A">
          ✓ No encontramos coincidencias significativas en las marcas registradas en Argentina.
        </p>

        <!-- CTA registro con descuento -->
        <div style="margin-top:28px;padding:22px;background:linear-gradient(135deg,#F0F5FF 0%,#fff 100%);
                    border:2px solid #1B6EF3;border-radius:14px;text-align:center">
          <div style="font-weight:700;font-size:17px;color:#0D1B4B;margin-bottom:6px">
            ¿Querés registrar
            <span x-text="'&quot;' + buscarResult.marca + '&quot;'" style="color:#1B6EF3"></span>?
          </div>
          <p style="color:#475569;margin:0 0 16px;font-size:14px">
            Como cliente Premium tenés
            <strong style="color:#16A34A">descuento exclusivo en honorarios</strong>
            de registro. Te contactamos por WhatsApp y armamos el trámite.
          </p>
          <a :href="ctaRegistroWa()" target="_blank" rel="noopener"
             style="display:inline-block;background:#25D366;color:#fff;text-decoration:none;
                    padding:14px 28px;border-radius:10px;font-weight:700;font-size:15px">
            💬 Iniciar registro por WhatsApp
          </a>
          <p style="margin:12px 0 0;font-size:12px;color:#64748b">
            Te respondemos en horario hábil (lun-vie 9-18hs ART).
          </p>
        </div>
      </div>
    </template>
  </div>

  <!-- CONSULTAS -->
  <div x-show="tab==='consultas'" class="card">
    <h3 style="margin-top:0">Historial de consultas</h3>
    <p x-show="!consultas.length" class="empty">Todavía no hiciste ninguna consulta. <a href="/">Buscar una marca →</a></p>
    <table x-show="consultas.length">
      <thead><tr><th>Marca</th><th>Nivel</th><th>Diagnóstico</th><th>Fecha</th><th></th></tr></thead>
      <tbody>
        <template x-for="c in consultas" :key="c.id">
          <tr>
            <td><strong x-text="c.marca"></strong></td>
            <td><span class="badge" :class="c.nivel==='completa'?'green':'gray'" x-text="c.nivel"></span></td>
            <td>
              <span class="badge" :class="diagBadge(c.diagnostico)" x-text="c.diagnostico||'pendiente'"></span>
            </td>
            <td x-text="fmtDate(c.created_at)"></td>
            <td>
              <button class="small sec" @click="abrirConsultaDetalle(c)">Ver informe →</button>
            </td>
          </tr>
        </template>
      </tbody>
    </table>
  </div>

  <!-- MIS MARCAS -->
  <div x-show="tab==='marcas'" class="card">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px;flex-wrap:wrap;gap:8px">
      <h3 style="margin:0">Mis marcas</h3>
      <div style="display:flex;gap:8px">
        <button class="small sec" @click="modalBulk=true">Cargar desde Excel</button>
        <button class="small" @click="modalMarca=true">+ Agregar marca</button>
      </div>
    </div>
    <p style="color:#64748b;font-size:14px;margin-top:0">
      Cargá las marcas que tenés registradas, en trámite o de terceros que te interesa seguir.
      Te avisamos las fechas clave: oposición (90 días desde la publicación),
      <strong>DJU</strong> a los 5 años y <strong>renovación</strong> a los 10.
    </p>
    <p x-show="!marcas.length" class="empty">Todavía no cargaste ninguna marca. Tocá "Agregar marca" o "Cargar desde Excel".</p>
    <div x-show="marcas.length" style="overflow-x:auto">
    <table>
      <thead><tr>
        <th>Marca</th><th>Clase</th><th>Tipo</th><th>Estado</th>
        <th>Oposición vence</th><th>DJU (5 años)</th><th>Vence (10 años)</th><th></th>
      </tr></thead>
      <tbody>
        <template x-for="m in marcas" :key="m.id">
          <tr>
            <td><strong x-text="m.denominacion"></strong>
              <div style="font-size:12px;color:#64748b">
                <span x-show="m.acta">Acta <span x-text="m.acta"></span></span>
                <span x-show="!m.es_propia && m.titular"> · titular: <span x-text="m.titular"></span></span>
              </div>
            </td>
            <td x-text="m.clase || '—'"></td>
            <td>
              <span class="badge" :class="m.es_propia ? 'green' : 'gray'"
                    x-text="m.es_propia ? 'Propia' : 'Tercero'"></span>
            </td>
            <td><span class="badge gray" x-text="m.estado||'—'"></span></td>
            <td>
              <span x-text="fmtDate(m.fecha_oposicion_fin)"></span>
              <span class="badge" :class="hitoBadge(m.fecha_oposicion_fin)"
                    x-show="m.fecha_oposicion_fin" x-text="hitoLabel(m.fecha_oposicion_fin)"></span>
            </td>
            <td>
              <span x-text="fmtDate(m.fecha_dju)"></span>
              <span class="badge" :class="hitoBadge(m.fecha_dju)"
                    x-show="m.fecha_dju" x-text="hitoLabel(m.fecha_dju)"></span>
            </td>
            <td>
              <span x-text="fmtDate(m.fecha_vencimiento)"></span>
              <span class="badge" :class="hitoBadge(m.fecha_vencimiento)"
                    x-show="m.fecha_vencimiento" x-text="hitoLabel(m.fecha_vencimiento)"></span>
            </td>
            <td style="white-space:nowrap">
              <button class="small sec" @click="abrirEditar(m)"
                      style="margin-right:4px">Editar</button>
              <button class="small sec" x-show="!hasVigilancia(m.id)"
                      @click="iniciarVigilancia(m.id)">Activar vigilancia</button>
              <span class="badge green" x-show="hasVigilancia(m.id)">Vigilando</span>
            </td>
          </tr>
        </template>
      </tbody>
    </table>
    </div>
  </div>

  <!-- VIGILANCIA -->
  <div x-show="tab==='vigilancia'" class="card">
    <h3 style="margin-top:0">Suscripciones de vigilancia</h3>
    <p style="color:#64748b">
      Cada semana, cuando el INPI publica su boletín, escaneamos automáticamente
      las marcas nuevas y te alertamos por email y en este panel si alguna se parece a las tuyas.
      Tu plan <strong>Premium</strong> incluye <strong x-text="precios.vigilancia_cap + ' vigilancias activas'"></strong>;
      las adicionales son $<span x-text="precios.vigilancia_marca.toLocaleString('es-AR')"></span> ARS / mes cada una.
    </p>
    <p x-show="!vigilancia.length" class="empty">
      Aún no tenés vigilancia activa. Andá a <a href="#" @click.prevent="tab='marcas'">Mis marcas</a> y activala desde una marca cargada.
    </p>
    <table x-show="vigilancia.length">
      <thead><tr><th>Marca vigilada</th><th>Tipo</th><th>Monto</th><th>Estado</th><th></th></tr></thead>
      <tbody>
        <template x-for="v in vigilancia" :key="v.id">
          <tr>
            <td x-text="v.marca_nombre || '(portfolio)'"></td>
            <td>
              <span x-text="v.tipo"></span>
              <span class="badge green" x-show="v.monto === 0">Incluida en Premium</span>
            </td>
            <td><span x-show="v.monto > 0">$<span x-text="v.monto.toLocaleString('es-AR')"></span></span>
                <span x-show="v.monto === 0" style="color:#16A34A">$0</span></td>
            <td>
              <span class="badge" :class="v.status==='active'?'green':'gray'" x-text="v.status"></span>
            </td>
            <td>
              <button class="small danger" x-show="v.status==='active'"
                      @click="cancelarVigilancia(v.id)">Cancelar</button>
            </td>
          </tr>
        </template>
      </tbody>
    </table>
  </div>

  <!-- ALERTAS -->
  <div x-show="tab==='alertas'" class="card">
    <h3 style="margin-top:0">Alertas detectadas</h3>
    <p x-show="!alertas.length" class="empty">No hay alertas. Si activás vigilancia, te avisamos acá y por email.</p>
    <template x-for="a in alertas" :key="a.id">
      <div class="alert-row" :class="a.nivel">
        <div>
          <strong x-text="a.marca_nueva_denominacion"></strong>
          <span style="color:#64748b;font-size:13px">— similar a tu marca <strong x-text="a.marca_propia"></strong></span>
          <div style="font-size:13px;margin-top:4px">
            Clase <span x-text="a.marca_nueva_clase"></span> ·
            Titular: <span x-text="a.marca_nueva_titular"></span> ·
            Score: <strong x-text="(a.score*100).toFixed(0)+'%'"></strong>
          </div>
        </div>
        <span class="badge" :class="a.nivel==='alto'?'red':a.nivel==='medio'?'yellow':'gray'"
              x-text="a.nivel" style="text-transform:uppercase"></span>
      </div>
    </template>
  </div>

  <!-- PAGOS -->
  <div x-show="tab==='pagos'" class="card">
    <h3 style="margin-top:0">Historial de pagos</h3>
    <p x-show="!pagos.length" class="empty">No hay pagos registrados todavía.</p>
    <table x-show="pagos.length">
      <thead><tr><th>Concepto</th><th>Monto</th><th>Estado</th><th>Fecha</th></tr></thead>
      <tbody>
        <template x-for="p in pagos" :key="p.id">
          <tr>
            <td x-text="p.tipo"></td>
            <td>$<span x-text="p.monto.toLocaleString('es-AR')"></span></td>
            <td><span class="badge" :class="p.status==='approved'?'green':p.status==='pending'?'yellow':'gray'" x-text="p.status"></span></td>
            <td x-text="fmtDate(p.created_at)"></td>
          </tr>
        </template>
      </tbody>
    </table>
  </div>

  <!-- MI PERFIL -->
  <div x-show="tab==='perfil'" class="card">
    <h3 style="margin-top:0">Mi perfil</h3>
    <p style="color:#64748b;margin:0 0 24px;font-size:14px">
      Cargá tus datos personales, redes y de tu empresa. Lo usamos para personalizar
      el servicio y, próximamente, para conectarte con otros emprendedores y agencias
      según tus necesidades y lo que ofrecés.
    </p>

    <div style="display:flex;gap:20px;align-items:center;margin-bottom:24px;padding:16px;
                background:#F4F5F9;border-radius:12px">
      <div style="position:relative">
        <div x-show="!perfil.avatar_data_uri"
             style="width:96px;height:96px;border-radius:50%;background:linear-gradient(135deg,#1B6EF3,#0D1B4B);
                    color:#fff;display:flex;align-items:center;justify-content:center;font-weight:700;font-size:32px;letter-spacing:1px"
             x-text="userInitials()"></div>
        <img x-show="perfil.avatar_data_uri" :src="perfil.avatar_data_uri" x-cloak
             style="width:96px;height:96px;border-radius:50%;object-fit:cover;border:3px solid #fff;box-shadow:0 2px 8px rgba(13,27,75,.15)">
      </div>
      <div style="flex:1">
        <div style="font-weight:700;font-size:18px" x-text="user.nombre || user.email.split('@')[0]"></div>
        <div style="font-size:14px;color:#64748b" x-text="user.email"></div>
        <div style="margin-top:10px;display:flex;gap:8px;flex-wrap:wrap">
          <label class="small sec" style="cursor:pointer;padding:8px 14px;border:1px solid #1B6EF3;background:#fff;color:#1B6EF3;border-radius:8px;font-weight:600;font-size:13px">
            <span x-text="perfil.avatar_data_uri ? 'Cambiar foto' : 'Subir foto'"></span>
            <input type="file" accept="image/*" style="display:none" @change="subirAvatar($event)">
          </label>
          <button x-show="perfil.avatar_data_uri" class="small danger"
                  @click="perfil.avatar_data_uri=''" style="padding:8px 14px">Quitar</button>
        </div>
        <p style="font-size:11px;color:#64748b;margin:8px 0 0">Recomendado: 200x200 px, máx 1 MB.</p>
      </div>
    </div>

    <h4 style="margin:0 0 12px;font-size:14px;color:#1B6EF3;text-transform:uppercase;letter-spacing:.5px">Datos personales</h4>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
      <div>
        <label>DNI / CUIT</label>
        <input type="text" x-model="perfil.dni">
      </div>
      <div>
        <label>Fecha de nacimiento</label>
        <input type="date" x-model="perfil.fecha_nacimiento">
      </div>
    </div>
    <label>Dirección</label>
    <input type="text" x-model="perfil.direccion" placeholder="Calle y número">
    <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px">
      <div>
        <label>Provincia</label>
        <select x-model="perfil.provincia" @change="perfil.localidad=''">
          <option value="">—</option>
          <template x-for="p in Object.keys(PROVINCIAS_AR)" :key="p">
            <option :value="p" x-text="p"></option>
          </template>
        </select>
      </div>
      <div>
        <label>Localidad</label>
        <input type="text" x-model="perfil.localidad" list="localidades-list"
               placeholder="Empezá a escribir...">
        <datalist id="localidades-list">
          <template x-for="loc in (PROVINCIAS_AR[perfil.provincia] || [])" :key="loc">
            <option :value="loc"></option>
          </template>
        </datalist>
      </div>
      <div><label>País</label><input type="text" x-model="perfil.pais"></div>
    </div>

    <h4 style="margin:28px 0 12px;font-size:14px;color:#1B6EF3;text-transform:uppercase;letter-spacing:.5px">Redes y web</h4>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
      <div><label>Sitio web</label><input type="url" x-model="perfil.web" placeholder="https://"></div>
      <div><label>Instagram</label><input type="text" x-model="perfil.instagram" placeholder="@usuario"></div>
      <div><label>LinkedIn</label><input type="text" x-model="perfil.linkedin" placeholder="linkedin.com/in/..."></div>
      <div><label>Twitter / X</label><input type="text" x-model="perfil.twitter" placeholder="@usuario"></div>
      <div><label>Facebook</label><input type="text" x-model="perfil.facebook"></div>
      <div><label>TikTok</label><input type="text" x-model="perfil.tiktok" placeholder="@usuario"></div>
    </div>

    <h4 style="margin:28px 0 12px;font-size:14px;color:#1B6EF3;text-transform:uppercase;letter-spacing:.5px">Empresa</h4>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
      <div><label>Nombre de la empresa</label><input type="text" x-model="perfil.empresa_nombre"></div>
      <div><label>Tu rol</label><input type="text" x-model="perfil.empresa_rol" placeholder="Founder, CMO, etc."></div>
      <div><label>Industria</label><input type="text" x-model="perfil.empresa_industria" placeholder="E-commerce, alimentos..."></div>
      <div>
        <label>Tamaño</label>
        <select x-model="perfil.empresa_tamano">
          <option value="">—</option>
          <option value="solo">Solo yo</option>
          <option value="2-10">2 a 10 personas</option>
          <option value="11-50">11 a 50 personas</option>
          <option value="51-200">51 a 200 personas</option>
          <option value="200+">+200 personas</option>
        </select>
      </div>
      <div><label>CUIT empresa</label><input type="text" x-model="perfil.empresa_cuit"></div>
    </div>

    <h4 style="margin:28px 0 12px;font-size:14px;color:#1B6EF3;text-transform:uppercase;letter-spacing:.5px">
      Tu proyecto <span style="color:#64748b;font-weight:400;font-size:12px;text-transform:none">(para futuras conexiones entre usuarios)</span>
    </h4>
    <label>Bio corta</label>
    <textarea x-model="perfil.bio" rows="2"
              placeholder="En una línea, ¿qué hacés?"
              style="width:100%;padding:10px;border:1px solid #E2E8F0;border-radius:8px;font-family:inherit;font-size:14px"></textarea>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
      <div>
        <label>Qué ofrezco</label>
        <textarea x-model="perfil.ofrece" rows="4"
                  placeholder="Ej: branding y diseño para PyMEs, software a medida, marketing digital..."
                  style="width:100%;padding:10px;border:1px solid #E2E8F0;border-radius:8px;font-family:inherit;font-size:14px"></textarea>
      </div>
      <div>
        <label>Qué busco</label>
        <textarea x-model="perfil.necesidades" rows="4"
                  placeholder="Ej: contador, programador freelance, agencia de marketing, inversor..."
                  style="width:100%;padding:10px;border:1px solid #E2E8F0;border-radius:8px;font-family:inherit;font-size:14px"></textarea>
      </div>
    </div>

    <div style="margin-top:24px">
      <button @click="guardarPerfil()" :disabled="perfilLoading">
        <span x-show="!perfilLoading">Guardar mi perfil</span>
        <span x-show="perfilLoading" x-cloak>Guardando…</span>
      </button>
      <span x-show="perfilMsg" x-text="perfilMsg" style="margin-left:12px;color:#16A34A;font-weight:600"></span>
    </div>
  </div>

  <!-- CONFIGURACIÓN -->
  <!-- NOTORIAS (admin) -->
  <div x-show="tab==='notorias' && user.is_admin" class="card">
    <div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px;margin-bottom:14px">
      <h3 style="margin:0">⭐ Lista de marcas notorias</h3>
      <div style="font-size:13px;color:#64748b">
        Total: <strong x-text="notoriasItems.length"></strong>
        · Default: <strong x-text="notoriasItems.filter(x => x.source === 'default').length"></strong>
        · Custom: <strong x-text="notoriasItems.filter(x => x.source === 'custom').length"></strong>
      </div>
    </div>
    <p style="color:#64748b;font-size:13px;margin:0 0 18px">
      Las marcas notorias tienen protección extendida a las 45 clases — el sistema
      las detecta como riesgo alto aunque el usuario busque en otra clase. Esta lista
      se aplica a todas las búsquedas del portal. <strong>Solo vos podés editarla</strong>.
    </p>

    <div style="display:flex;gap:8px;margin-bottom:14px;flex-wrap:wrap">
      <input type="text" x-model="notoriasFiltro" placeholder="🔎 Filtrar..."
             autocomplete="off" style="flex:1;min-width:200px">
      <input type="text" x-model="notoriasNueva" @keydown.enter="agregarNotoria()"
             placeholder="Agregar marca nueva (Enter para sumar)"
             autocomplete="off" style="flex:2;min-width:200px">
      <button @click="agregarNotoria()" :disabled="!notoriasNueva.trim()">Agregar</button>
    </div>

    <div style="max-height:60vh;overflow-y:auto;border:1px solid #E2E8F0;border-radius:8px">
      <table style="margin:0">
        <thead><tr>
          <th>Marca</th><th>Origen</th><th></th>
        </tr></thead>
        <tbody>
          <template x-for="b in notoriasFiltradas()" :key="b.denominacion">
            <tr>
              <td><strong x-text="b.denominacion"></strong></td>
              <td>
                <span class="badge" :class="b.source === 'default' ? 'gray' : 'green'"
                      x-text="b.source === 'default' ? 'sistema' : 'agregada'"></span>
              </td>
              <td style="text-align:right">
                <button class="small danger" @click="eliminarNotoria(b.denominacion)">
                  Eliminar
                </button>
              </td>
            </tr>
          </template>
        </tbody>
      </table>
      <p x-show="!notoriasFiltradas().length" class="empty" style="padding:30px">
        Sin resultados — probá con otro filtro.
      </p>
    </div>
  </div>

  <div x-show="tab==='config'" class="card">
    <h3 style="margin-top:0">Configuración de cuenta</h3>

    <label>Nombre</label>
    <input type="text" x-model="config.nombre" placeholder="Cómo querés que te llamemos">

    <label>Teléfono</label>
    <div style="display:grid;grid-template-columns:220px 1fr;gap:8px">
      <select x-model="config.tel_cc">
        <template x-for="c in COUNTRY_CODES" :key="c.iso">
          <option :value="c.code" x-text="c.flag + ' ' + c.name + ' (+' + c.code + ')'"></option>
        </template>
      </select>
      <input type="tel" x-model="config.tel_num"
             :placeholder="config.tel_cc === '54' ? '9 11 1234-5678' : 'Número (sin el código de país)'">
    </div>

    <div style="margin:18px 0;padding:14px;background:#F4F5F9;border-radius:10px">
      <label style="display:flex;gap:10px;align-items:center;cursor:pointer;font-weight:600">
        <input type="checkbox" x-model="config.alertas_whatsapp" style="width:auto">
        Recibir alertas también por WhatsApp
      </label>
      <p style="font-size:13px;color:#64748b;margin:8px 0 0">
        Te avisamos por WhatsApp además del email cuando aparezca una marca similar a las
        tuyas o se acerque un vencimiento. Necesitás tener tu teléfono cargado arriba.
      </p>
    </div>

    <button @click="guardarConfig()" :disabled="configLoading">
      <span x-show="!configLoading">Guardar cambios</span>
      <span x-show="configLoading" x-cloak>Guardando…</span>
    </button>
    <span x-show="configMsg" x-text="configMsg" style="margin-left:12px;color:#16A34A;font-weight:600"></span>

    <hr style="border:none;border-top:1px solid #E2E8F0;margin:32px 0">
    <h3>Tu suscripción</h3>
    <div x-show="premium" style="background:#F0F5FF;padding:16px;border-radius:10px">
      <p style="margin:0 0 6px"><strong x-text="premium ? ('Plan ' + (premium.plan_freq||'mensual')) : ''"></strong>
         · $<span x-text="(premium?.monto||0).toLocaleString('es-AR')"></span>
         <span x-text="premium?.plan_freq === 'anual' ? '/ año' : '/ mes'"></span></p>
      <p style="margin:0;color:#64748b;font-size:14px" x-show="premium?.paid_through_date">
        Vigente hasta <strong x-text="fmtDate(premium?.paid_through_date)"></strong>
      </p>

      <label style="display:flex;gap:10px;align-items:flex-start;cursor:pointer;
                    margin-top:14px;padding-top:14px;border-top:1px solid #DBEAFE">
        <input type="checkbox" x-model="premiumRenew" @change="cambiarAutoRenew()"
               style="width:auto;margin-top:3px">
        <div>
          <div style="font-weight:600">Renovación automática</div>
          <div style="font-size:13px;color:#64748b">
            Si lo dejás activado, te cobramos automáticamente al fin de cada período.
            Si lo desactivás, te avisamos antes del vencimiento para que renueves vos.
          </div>
        </div>
      </label>
    </div>
    <p x-show="!premium" style="color:#64748b">No tenés información de tu suscripción.</p>
  </div>

  <!-- MODAL agregar / editar marca -->
  <div class="modal" x-show="modalMarca" x-cloak @click.self="cerrarModalMarca()">
    <div class="modal-content" style="max-width:560px">
      <h3 style="margin-top:0" x-text="editandoMarcaId ? 'Editar marca' : 'Agregar marca'"></h3>

      <label>Denominación *</label>
      <input type="text" x-model="nueva.denominacion" placeholder="Mi Marca">

      <label>Tipo</label>
      <select x-model="nueva.es_propia">
        <option :value="true">Propia (la tengo o la voy a registrar)</option>
        <option :value="false">De tercero (la quiero monitorear)</option>
      </select>

      <label x-show="!nueva.es_propia">Titular (de quién es)</label>
      <input x-show="!nueva.es_propia" type="text" x-model="nueva.titular"
             placeholder="Ej: Acme S.A.">

      <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
        <div>
          <label>Clase Niza</label>
          <input type="number" x-model.number="nueva.clase" min="1" max="45" placeholder="ej: 25">
        </div>
        <div>
          <label>Acta INPI (opcional)</label>
          <input type="text" x-model="nueva.acta">
        </div>
      </div>

      <label>Estado (opcional)</label>
      <select x-model="nueva.estado">
        <option value="">—</option>
        <option value="solicitada">Solicitada (en trámite)</option>
        <option value="publicada">Publicada en boletín</option>
        <option value="oposicion">Con oposición</option>
        <option value="concedida">Concedida / Registrada</option>
        <option value="vencida">Vencida</option>
      </select>

      <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
        <div>
          <label>Inicio de trámite</label>
          <input type="date" x-model="nueva.fecha_solicitud">
        </div>
        <div>
          <label>Fecha de publicación</label>
          <input type="date" x-model="nueva.fecha_publicacion">
          <p style="font-size:11px;color:#64748b;margin:2px 0 0">Para contar los 90 días de oposición.</p>
        </div>
      </div>

      <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
        <div>
          <label>Fecha de oposición</label>
          <input type="date" x-model="nueva.fecha_oposicion">
          <p style="font-size:11px;color:#64748b;margin:2px 0 0">Si hubo oposición.</p>
        </div>
        <div>
          <label>Fecha de concesión</label>
          <input type="date" x-model="nueva.fecha_concesion">
          <p style="font-size:11px;color:#64748b;margin:2px 0 0">Si la cargás, calculamos DJU y vencimiento.</p>
        </div>
      </div>

      <div style="display:flex;gap:8px;margin-top:20px">
        <button class="sec" @click="cerrarModalMarca()" style="flex:1">Cancelar</button>
        <button @click="guardarMarca()" style="flex:1">
          <span x-text="editandoMarcaId ? 'Guardar cambios' : 'Crear marca'"></span>
        </button>
      </div>
    </div>
  </div>

  <!-- MODAL detalle de consulta -->
  <div class="modal" x-show="modalConsulta" x-cloak @click.self="modalConsulta=false">
    <div class="modal-content" style="max-width:780px;max-height:85vh;overflow-y:auto">
      <h3 style="margin-top:0" x-text="'Informe — ' + (consultaDetalle?.marca || '')"></h3>
      <p style="color:#64748b;font-size:13px;margin:0 0 16px">
        Creada el <span x-text="fmtDate(consultaDetalle?.created_at)"></span>
        · <span x-text="consultaDetalle?.nivel"></span>
      </p>

      <div x-show="consultaCargando" class="empty">Cargando informe...</div>

      <div x-show="!consultaCargando && consultaDetalle">
        <div x-show="consultaDetalle?.diagnostico" style="margin-bottom:16px">
          <span class="badge" :class="diagBadge(consultaDetalle?.diagnostico)"
                x-text="consultaDetalle?.diagnostico"></span>
        </div>

        <div x-show="consultaDetalle?.pre_analisis_ia"
             style="background:#F4F5F9;border-radius:10px;padding:16px;
                    font-size:14px;line-height:1.6;white-space:pre-wrap"
             x-text="consultaDetalle?.pre_analisis_ia"></div>

        <h4 style="margin:18px 0 8px">Coincidencias</h4>
        <p x-show="!consultaDetalle?.resultados || !consultaDetalle.resultados.length"
           class="empty" style="padding:20px">
          No se encontraron coincidencias significativas.
        </p>
        <table x-show="consultaDetalle?.resultados && consultaDetalle.resultados.length">
          <thead><tr>
            <th>Marca</th><th>Clase</th><th>Titular</th><th>Score</th><th>Nivel</th>
          </tr></thead>
          <tbody>
            <template x-for="r in (consultaDetalle?.resultados || [])" :key="r.id || r.denominacion">
              <tr>
                <td><strong x-text="r.denominacion"></strong></td>
                <td x-text="r.clase || '—'"></td>
                <td x-text="r.titular || '—'"></td>
                <td><strong x-text="((r.score||0)*100).toFixed(0)+'%'"></strong></td>
                <td>
                  <span class="badge" :class="{red:r.nivel==='alto',yellow:r.nivel==='medio',gray:r.nivel==='bajo'}"
                        x-text="r.nivel"></span>
                </td>
              </tr>
            </template>
          </tbody>
        </table>
      </div>

      <div style="display:flex;gap:8px;margin-top:24px">
        <button class="sec" @click="modalConsulta=false" style="flex:1">Cerrar</button>
      </div>
    </div>
  </div>

  <!-- MODAL cambiar contraseña -->
  <div class="modal" x-show="modalPassword" x-cloak @click.self="modalPassword=false">
    <div class="modal-content" style="max-width:420px">
      <h3 style="margin-top:0">Cambiar contraseña</h3>
      <label>Contraseña actual</label>
      <input type="password" x-model="pwForm.actual" autocomplete="current-password">
      <label>Contraseña nueva</label>
      <input type="password" x-model="pwForm.nueva" autocomplete="new-password">
      <p style="font-size:12px;color:#64748b;margin:4px 0 0">Mínimo 8 caracteres.</p>
      <label>Repetí la nueva</label>
      <input type="password" x-model="pwForm.repeat" autocomplete="new-password">
      <p x-show="pwError" x-text="pwError" style="color:#DC2626;font-size:13px;margin-top:8px"></p>
      <p x-show="pwOk" x-text="pwOk" style="color:#16A34A;font-size:13px;margin-top:8px"></p>
      <div style="display:flex;gap:8px;margin-top:20px">
        <button class="sec" @click="modalPassword=false; pwError=''; pwOk=''" style="flex:1">Cerrar</button>
        <button @click="cambiarPassword()" :disabled="pwLoading" style="flex:1">
          <span x-show="!pwLoading">Cambiar</span>
          <span x-show="pwLoading" x-cloak>Cambiando…</span>
        </button>
      </div>
    </div>
  </div>

  <!-- MODAL carga masiva (Excel/CSV) -->
  <div class="modal" x-show="modalBulk" x-cloak @click.self="modalBulk=false">
    <div class="modal-content" style="max-width:520px">
      <h3 style="margin-top:0">Cargar marcas desde Excel</h3>
      <p style="color:#475569;font-size:14px;margin-top:0">
        Subí un .xlsx o .csv con una fila por marca. La primera fila tiene que tener los nombres de las columnas.
      </p>

      <div style="background:#F0F5FF;border:1px solid #DBEAFE;border-radius:10px;
                  padding:14px;margin:14px 0">
        <div style="font-weight:600;color:#0D1B4B;margin-bottom:4px">¿No sabés cómo armarlo?</div>
        <p style="font-size:13px;color:#475569;margin:0 0 10px">
          Bajate nuestro template con los encabezados y filas de ejemplo. Lo completás
          en Excel y volvés a subirlo acá.
        </p>
        <a href="/api/dashboard/marcas/template" download
           class="sec" style="display:inline-block;padding:10px 16px;border:1px solid #1B6EF3;
                              background:#fff;color:#1B6EF3;text-decoration:none;
                              border-radius:8px;font-weight:600;font-size:14px">
          ↓ Descargar template .xlsx
        </a>
      </div>

      <p style="font-size:13px;color:#64748b">
        <strong>Columnas reconocidas:</strong>
        denominacion (obligatoria), clase, acta, titular, estado, es_propia (sí/no),
        fecha_solicitud, fecha_publicacion, fecha_oposicion,
        fecha_concesion, fecha_vencimiento, notas.
      </p>
      <input type="file" accept=".xlsx,.xlsm,.csv" @change="bulkFile = $event.target.files[0]">
      <div x-show="bulkResult" style="margin-top:14px;background:#F4F5F9;
                                       padding:12px;border-radius:8px;font-size:14px">
        <div><strong x-text="bulkResult?.creadas"></strong> marcas creadas.</div>
        <div x-show="bulkResult?.errores?.length" style="margin-top:8px">
          <strong>Errores:</strong>
          <ul style="margin:6px 0 0;padding-left:20px;color:#991B1B">
            <template x-for="e in (bulkResult?.errores || [])">
              <li><span x-text="'Fila ' + e.fila + ': ' + e.motivo"></span></li>
            </template>
          </ul>
        </div>
      </div>
      <div style="display:flex;gap:8px;margin-top:20px">
        <button class="sec" @click="modalBulk=false; bulkFile=null; bulkResult=null" style="flex:1">Cerrar</button>
        <button @click="subirBulk()" :disabled="!bulkFile || bulkLoading" style="flex:1">
          <span x-show="!bulkLoading">Subir archivo</span>
          <span x-show="bulkLoading" x-cloak>Procesando…</span>
        </button>
      </div>
    </div>
  </div>

</main>

<script>
function dashboard(){
  return {
    tab: new URLSearchParams(location.search).get('tab') || 'buscar',
    user: {email:'', nombre:'', is_admin:false},
    userMenuOpen: false,
    modalPassword: false,
    pwForm: {actual:'', nueva:'', repeat:''},
    pwError: '', pwOk: '', pwLoading: false,
    PROVINCIAS_AR: {
      "CABA": ["Capital Federal"],
      "Buenos Aires": ["La Plata","Mar del Plata","Bahía Blanca","Tandil","Quilmes","Lomas de Zamora","San Isidro","Tigre","Pilar","Morón","Tres de Febrero","La Matanza","Avellaneda","Lanús","Berazategui","Florencio Varela","Necochea","Olavarría","Pergamino","Junín","Luján","Campana","Zárate","San Nicolás"],
      "Catamarca": ["San Fernando del Valle de Catamarca","Belén","Andalgalá","Tinogasta"],
      "Chaco": ["Resistencia","Presidencia Roque Sáenz Peña","Villa Ángela","Charata"],
      "Chubut": ["Rawson","Comodoro Rivadavia","Trelew","Puerto Madryn","Esquel"],
      "Córdoba": ["Córdoba","Río Cuarto","Villa María","San Francisco","Carlos Paz","Alta Gracia","Jesús María","Bell Ville"],
      "Corrientes": ["Corrientes","Goya","Mercedes","Curuzú Cuatiá","Paso de los Libres"],
      "Entre Ríos": ["Paraná","Concordia","Gualeguaychú","Concepción del Uruguay","Gualeguay","Victoria"],
      "Formosa": ["Formosa","Clorinda","Pirané","Las Lomitas"],
      "Jujuy": ["San Salvador de Jujuy","Palpalá","San Pedro","Libertador General San Martín","La Quiaca"],
      "La Pampa": ["Santa Rosa","General Pico","Toay","Realicó"],
      "La Rioja": ["La Rioja","Chilecito","Aimogasta","Chamical"],
      "Mendoza": ["Mendoza","San Rafael","Godoy Cruz","Maipú","Luján de Cuyo","Las Heras","Tunuyán","General Alvear"],
      "Misiones": ["Posadas","Oberá","Eldorado","Puerto Iguazú","San Vicente","Apóstoles"],
      "Neuquén": ["Neuquén","Cutral Có","Plottier","Centenario","Zapala","San Martín de los Andes"],
      "Río Negro": ["Viedma","San Carlos de Bariloche","General Roca","Cipolletti","Cinco Saltos","Villa Regina"],
      "Salta": ["Salta","San Ramón de la Nueva Orán","Tartagal","Cafayate","Metán"],
      "San Juan": ["San Juan","Rivadavia","Caucete","Pocito","Santa Lucía"],
      "San Luis": ["San Luis","Villa Mercedes","Merlo","La Toma"],
      "Santa Cruz": ["Río Gallegos","Caleta Olivia","El Calafate","Pico Truncado","Puerto Deseado"],
      "Santa Fe": ["Santa Fe","Rosario","Rafaela","Reconquista","Venado Tuerto","Esperanza","Villa Constitución"],
      "Santiago del Estero": ["Santiago del Estero","La Banda","Termas de Río Hondo","Frías"],
      "Tierra del Fuego": ["Ushuaia","Río Grande","Tolhuin"],
      "Tucumán": ["San Miguel de Tucumán","Yerba Buena","Tafí Viejo","Concepción","Banda del Río Salí","Aguilares"],
    },
    NIZA_TITLES: {
      1:"Productos químicos",2:"Pinturas y barnices",3:"Cosméticos y limpieza",
      4:"Aceites y combustibles",5:"Farmacéuticos",6:"Metales comunes",
      7:"Máquinas y motores",8:"Herramientas de mano",9:"Electrónica y software",
      10:"Aparatos médicos",11:"Alumbrado y calefacción",12:"Vehículos",
      13:"Armas de fuego",14:"Joyería y relojes",15:"Instrumentos musicales",
      16:"Papel e imprenta",17:"Caucho y plásticos",18:"Cuero y bolsos",
      19:"Materiales de construcción",20:"Muebles",21:"Utensilios domésticos",
      22:"Cuerdas y fibras",23:"Hilos textiles",24:"Telas y textiles",
      25:"Ropa y calzado",26:"Encajes y bordados",27:"Alfombras",
      28:"Juegos y juguetes",29:"Carne y alimentos",30:"Café, té y panadería",
      31:"Agrícola y animales vivos",32:"Cervezas y bebidas",33:"Bebidas alcohólicas",
      34:"Tabaco",35:"Publicidad y gestión",36:"Seguros y finanzas",
      37:"Construcción y reparación",38:"Telecomunicaciones",39:"Transporte",
      40:"Tratamiento de materiales",41:"Educación y entretenimiento",
      42:"Servicios tecnológicos",43:"Restauración y alojamiento",
      44:"Servicios médicos y veterinarios",45:"Servicios jurídicos y seguridad"
    },
    buscar: {marca:'', clases:[], claseFiltro:'', descripcion:''},
    buscando: false, buscarErr: '', buscarResult: null,
    ayudaScores: false,
    modalConsulta: false, consultaDetalle: null, consultaCargando: false,
    notoriasItems: [], notoriasFiltro: '', notoriasNueva: '',
    consultas: [], marcas: [], vigilancia: [], alertas: [], pagos: [],
    precios: {vigilancia_marca: 1500, vigilancia_portfolio: 50000, vigilancia_cap: 10},
    modalMarca: false,
    editandoMarcaId: null,
    nueva: {
      denominacion:'', clase:null, acta:'', estado:'',
      es_propia: true, titular:'',
      fecha_solicitud:'', fecha_publicacion:'', fecha_oposicion:'', fecha_concesion:'',
    },
    modalBulk: false, bulkFile: null, bulkResult: null, bulkLoading: false,
    COUNTRY_CODES: [
      {iso:'AR', code:'54', name:'Argentina', flag:'🇦🇷'},
      {iso:'UY', code:'598', name:'Uruguay', flag:'🇺🇾'},
      {iso:'CL', code:'56', name:'Chile', flag:'🇨🇱'},
      {iso:'BR', code:'55', name:'Brasil', flag:'🇧🇷'},
      {iso:'PY', code:'595', name:'Paraguay', flag:'🇵🇾'},
      {iso:'BO', code:'591', name:'Bolivia', flag:'🇧🇴'},
      {iso:'PE', code:'51', name:'Perú', flag:'🇵🇪'},
      {iso:'CO', code:'57', name:'Colombia', flag:'🇨🇴'},
      {iso:'VE', code:'58', name:'Venezuela', flag:'🇻🇪'},
      {iso:'EC', code:'593', name:'Ecuador', flag:'🇪🇨'},
      {iso:'MX', code:'52', name:'México', flag:'🇲🇽'},
      {iso:'US', code:'1', name:'Estados Unidos', flag:'🇺🇸'},
      {iso:'ES', code:'34', name:'España', flag:'🇪🇸'},
      {iso:'IT', code:'39', name:'Italia', flag:'🇮🇹'},
      {iso:'FR', code:'33', name:'Francia', flag:'🇫🇷'},
      {iso:'DE', code:'49', name:'Alemania', flag:'🇩🇪'},
      {iso:'UK', code:'44', name:'Reino Unido', flag:'🇬🇧'},
    ],
    config: {nombre:'', tel_cc:'54', tel_num:'', alertas_whatsapp:false},
    configLoading: false, configMsg: '',
    perfil: {
      dni:'', fecha_nacimiento:'', direccion:'', localidad:'', provincia:'', pais:'Argentina',
      instagram:'', linkedin:'', twitter:'', facebook:'', tiktok:'', web:'',
      empresa_nombre:'', empresa_rol:'', empresa_industria:'', empresa_tamano:'', empresa_cuit:'',
      bio:'', ofrece:'', necesidades:'', avatar_data_uri:'',
    },
    perfilLoading: false, perfilMsg: '',
    premium: null, premiumRenew: true,

    async cargar(){
      const me = await fetch('/api/auth/me').then(r=>r.json());
      if(!me.data.authenticated){ location.href='/login'; return; }
      this.user = me.data;
      this.config.nombre = me.data.nombre || '';
      // Parsear teléfono guardado (formato E.164: +54...) en cc + num
      const tel = (me.data.telefono || '').trim();
      if (tel.startsWith('+')) {
        const match = this.COUNTRY_CODES
          .slice().sort((a,b) => b.code.length - a.code.length)
          .find(c => tel.slice(1).startsWith(c.code));
        if (match) {
          this.config.tel_cc = match.code;
          this.config.tel_num = tel.slice(1 + match.code.length);
        } else {
          this.config.tel_num = tel;
        }
      } else if (tel) {
        this.config.tel_num = tel;
      }
      this.config.alertas_whatsapp = !!me.data.alertas_whatsapp;
      await Promise.all([
        this.fetchConsultas(), this.fetchMarcas(),
        this.fetchVigilancia(), this.fetchAlertas(), this.fetchPagos(),
        this.fetchPrecios(), this.fetchPremium(), this.fetchPerfil(),
      ]);
    },
    async fetchPerfil(){
      const r = await fetch('/api/dashboard/perfil').then(r=>r.json());
      if (r.ok && r.data) {
        Object.assign(this.perfil, r.data);
      }
    },
    async guardarPerfil(){
      this.perfilLoading = true;
      this.perfilMsg = '';
      try {
        const r = await fetch('/api/dashboard/perfil', {
          method:'POST', headers:{'Content-Type':'application/json'},
          body: JSON.stringify(this.perfil),
        }).then(r=>r.json());
        if (!r.ok) { alert(r.error || 'Error guardando'); return; }
        this.perfilMsg = 'Perfil guardado ✓';
        setTimeout(() => this.perfilMsg = '', 3000);
      } finally {
        this.perfilLoading = false;
      }
    },
    async subirAvatar(evt){
      const file = evt.target.files && evt.target.files[0];
      if (!file) return;
      if (file.size > 1_500_000) {
        alert('La foto pesa más de 1.5 MB. Subí una más liviana.');
        evt.target.value = '';
        return;
      }
      const reader = new FileReader();
      reader.onload = () => {
        // Redimensionamos en canvas a 240x240 para no guardar imágenes enormes
        const img = new Image();
        img.onload = () => {
          const canvas = document.createElement('canvas');
          canvas.width = 240; canvas.height = 240;
          const ctx = canvas.getContext('2d');
          const min = Math.min(img.width, img.height);
          const sx = (img.width - min)/2, sy = (img.height - min)/2;
          ctx.drawImage(img, sx, sy, min, min, 0, 0, 240, 240);
          this.perfil.avatar_data_uri = canvas.toDataURL('image/jpeg', 0.85);
        };
        img.src = reader.result;
      };
      reader.readAsDataURL(file);
      evt.target.value = '';
    },
    async fetchPremium(){
      const r = await fetch('/api/dashboard/premium').then(r=>r.json());
      this.premium = r.data || null;
      this.premiumRenew = this.premium ? !!this.premium.auto_renew : true;
    },
    async cambiarAutoRenew(){
      const r = await fetch('/api/dashboard/premium/auto_renew', {
        method:'POST', headers:{'Content-Type':'application/json'},
        body: JSON.stringify({auto_renew: this.premiumRenew}),
      }).then(r=>r.json());
      if(!r.ok){ alert(r.error || 'Error'); this.premiumRenew = !this.premiumRenew; return; }
      this.configMsg = this.premiumRenew ? 'Renovación automática activada ✓' : 'Renovación automática desactivada ✓';
      setTimeout(() => this.configMsg = '', 4000);
      await this.fetchPremium();
    },
    async guardarConfig(){
      this.configLoading = true;
      this.configMsg = '';
      const telefonoFull = this.config.tel_num.trim()
        ? '+' + this.config.tel_cc + this.config.tel_num.replace(/[^\d]/g,'')
        : '';
      try {
        const r = await fetch('/api/dashboard/config', {
          method:'POST', headers:{'Content-Type':'application/json'},
          body: JSON.stringify({
            nombre: this.config.nombre,
            telefono: telefonoFull,
            alertas_whatsapp: this.config.alertas_whatsapp,
          }),
        }).then(r=>r.json());
        if(!r.ok){ alert(r.error||'Error guardando'); return; }
        this.configMsg = 'Guardado ✓';
        setTimeout(() => this.configMsg = '', 3000);
      } finally {
        this.configLoading = false;
      }
    },
    async fetchConsultas(){ this.consultas = (await fetch('/api/dashboard/consultas').then(r=>r.json())).data || []; },
    async fetchMarcas(){    this.marcas    = (await fetch('/api/dashboard/marcas').then(r=>r.json())).data || []; },
    async fetchVigilancia(){this.vigilancia= (await fetch('/api/dashboard/vigilancia').then(r=>r.json())).data || []; },
    async fetchAlertas(){   this.alertas   = (await fetch('/api/dashboard/alertas').then(r=>r.json())).data || []; },
    async fetchPagos(){     this.pagos     = (await fetch('/api/dashboard/pagos').then(r=>r.json())).data || []; },
    async fetchPrecios(){   this.precios   = (await fetch('/api/dashboard/precios').then(r=>r.json())).data || this.precios; },

    userInitials(){
      const src = (this.user.nombre || this.user.email || 'U').trim();
      const parts = src.split(/[\s.@_-]+/).filter(Boolean);
      if (parts.length >= 2) return (parts[0][0] + parts[1][0]).toUpperCase();
      return src.slice(0, 2).toUpperCase();
    },

    async cambiarPassword(){
      this.pwError = ''; this.pwOk = '';
      if (!this.pwForm.actual || !this.pwForm.nueva) {
        this.pwError = 'Completá ambas contraseñas.'; return;
      }
      if (this.pwForm.nueva.length < 8) {
        this.pwError = 'La nueva contraseña tiene que tener al menos 8 caracteres.'; return;
      }
      if (this.pwForm.nueva !== this.pwForm.repeat) {
        this.pwError = 'La nueva no coincide con la repetición.'; return;
      }
      this.pwLoading = true;
      try {
        const r = await fetch('/api/auth/change-password', {
          method:'POST', headers:{'Content-Type':'application/json'},
          body: JSON.stringify({actual: this.pwForm.actual, nueva: this.pwForm.nueva}),
        }).then(r=>r.json());
        if (!r.ok) {
          this.pwError = r.error || 'No pudimos cambiar la contraseña.';
        } else {
          this.pwOk = 'Contraseña actualizada ✓';
          this.pwForm = {actual:'', nueva:'', repeat:''};
          setTimeout(() => { this.modalPassword = false; this.pwOk = ''; }, 1500);
        }
      } finally {
        this.pwLoading = false;
      }
    },

    async abrirConsultaDetalle(c){
      this.modalConsulta = true;
      this.consultaDetalle = {marca: c.marca, created_at: c.created_at,
                              nivel: c.nivel, diagnostico: c.diagnostico};
      this.consultaCargando = true;
      try {
        const r = await fetch('/api/marca/consulta/' + c.id).then(r=>r.json());
        if (r.ok) this.consultaDetalle = r.data;
      } finally {
        this.consultaCargando = false;
      }
    },

    matchTags(m, query, clase){
      const tags = [];
      const normalize = s => (s||'').toLowerCase().normalize('NFD').replace(/[̀-ͯ]/g,'').trim();
      const q = normalize(query);
      const d = normalize(m.denominacion);
      if (q && d && q === d) tags.push('idéntica');
      else if (q && d && (d.startsWith(q) || q.startsWith(d))) tags.push('misma raíz');
      const scores = m.scores || {};
      if ((scores.fonetica||0) >= 0.85) tags.push('suena igual');
      if ((scores.conceptual||0) >= 0.75) tags.push('mismo concepto');
      if ((m.estado_code||'').toLowerCase() === 'vigente') tags.push('vigente');
      return tags;
    },

    async fetchNotorias(){
      const r = await fetch('/api/marca/notorias').then(r=>r.json());
      if (r.ok) this.notoriasItems = r.data.items || [];
    },
    notoriasFiltradas(){
      const q = (this.notoriasFiltro || '').toLowerCase();
      if (!q) return this.notoriasItems;
      return this.notoriasItems.filter(b =>
        b.denominacion.toLowerCase().includes(q));
    },
    async agregarNotoria(){
      const name = (this.notoriasNueva || '').trim();
      if (!name) return;
      const r = await fetch('/api/marca/notorias/agregar', {
        method:'POST', headers:{'Content-Type':'application/json'},
        body: JSON.stringify({denominacion: name}),
      }).then(r=>r.json());
      if (!r.ok) { alert(r.error || 'Error'); return; }
      this.notoriasNueva = '';
      await this.fetchNotorias();
    },
    async eliminarNotoria(name){
      if (!confirm(`¿Quitar "${name}" de la lista de notorias?`)) return;
      const r = await fetch('/api/marca/notorias/eliminar', {
        method:'POST', headers:{'Content-Type':'application/json'},
        body: JSON.stringify({denominacion: name}),
      }).then(r=>r.json());
      if (!r.ok) { alert(r.error || 'Error'); return; }
      await this.fetchNotorias();
    },

    async marcarTodasComoNotorias(){
      const altos = (this.buscarResult?.matches || []).filter(m => m.nivel === 'alto');
      if (!altos.length) { alert('No hay coincidencias nivel alto para marcar.'); return; }
      if (!confirm(`Marcar ${altos.length} marca(s) como notorias?\n\n` +
                    altos.slice(0,5).map(m => '• ' + m.denominacion).join('\\n') +
                    (altos.length > 5 ? `\\n• ... y ${altos.length - 5} más` : ''))) return;
      let ok = 0;
      for (const m of altos) {
        const r = await fetch('/api/marca/notorias/agregar', {
          method:'POST', headers:{'Content-Type':'application/json'},
          body: JSON.stringify({denominacion: m.denominacion}),
        }).then(r=>r.json());
        if (r.ok && r.data.added) ok += 1;
      }
      alert(`${ok} marca(s) agregada(s) a la lista de notorias.`);
    },

    ctaRegistroWa(){
      const marca = this.buscarResult?.marca || '';
      const clases = this.buscarResult?.clases_consultadas || [];
      const clasesTxt = clases.length ? ` en la(s) clase(s) ${clases.join(', ')}` : '';
      const veredicto = this.buscarResult?.veredicto || '';
      const verTxt = veredicto === 'no_disponible'
        ? ' El sistema marcó riesgo alto, pero quiero asesorarme.'
        : (veredicto === 'necesita_analisis' ? ' Necesito un análisis más profundo.' : '');
      const msg = `Hola LegalPacers, soy cliente Premium y quiero iniciar el registro de la marca "${marca}"${clasesTxt}.${verTxt} Por favor coordinen el descuento exclusivo.`;
      return `https://wa.me/5491128774200?text=${encodeURIComponent(msg)}`;
    },

    async ejecutarBusqueda(){
      if(!this.buscar.marca.trim()){
        this.buscarErr = 'Ingresá una marca para buscar.';
        return;
      }
      this.buscarErr = '';
      this.buscarResult = null;
      this.buscando = true;
      try {
        const r = await fetch('/api/marca/check', {
          method:'POST', headers:{'Content-Type':'application/json'},
          body: JSON.stringify({
            marca: this.buscar.marca,
            descripcion: this.buscar.descripcion,
            clases: (this.buscar.clases || []).map(c => parseInt(c)).filter(Boolean),
          }),
        });
        const d = await r.json();
        if(!d.ok){
          this.buscarErr = d.error || 'No pudimos procesar la búsqueda.';
        } else {
          this.buscarResult = d.data;
          await this.fetchConsultas();
        }
      } catch(e){
        this.buscarErr = 'Error de red.';
      } finally {
        this.buscando = false;
      }
    },

    fmtDate(d){ return d ? new Date(d).toLocaleDateString('es-AR') : '—'; },
    diagBadge(d){ return ({
      'viable':'green','viable_con_ajustes':'yellow','riesgo_alto':'red',
    })[d] || 'gray'; },
    _daysTo(dateStr){
      if(!dateStr) return null;
      const target = new Date(dateStr);
      const now = new Date();
      return Math.floor((target - now) / (1000*60*60*24));
    },
    hitoBadge(dateStr){
      const d = this._daysTo(dateStr);
      if(d === null) return 'gray';
      if(d < 0) return 'red';
      if(d <= 30) return 'red';
      if(d <= 90) return 'yellow';
      return 'green';
    },
    hitoLabel(dateStr){
      const d = this._daysTo(dateStr);
      if(d === null) return '';
      if(d < 0) return 'vencido';
      if(d === 0) return 'hoy';
      if(d <= 30) return d + 'd';
      if(d <= 90) return d + 'd';
      if(d <= 365) return Math.round(d/30) + 'm';
      return Math.round(d/365) + 'a';
    },
    hasVigilancia(marcaId){
      return this.vigilancia.some(v => v.marca_cliente_id===marcaId && v.status==='active');
    },
    abrirEditar(m){
      this.editandoMarcaId = m.id;
      this.nueva = {
        denominacion: m.denominacion || '',
        clase: m.clase || null,
        acta: m.acta || '',
        estado: m.estado || '',
        es_propia: m.es_propia !== false,
        titular: m.titular || '',
        fecha_solicitud: m.fecha_solicitud || '',
        fecha_publicacion: m.fecha_publicacion || '',
        fecha_oposicion: m.fecha_oposicion || '',
        fecha_concesion: m.fecha_concesion || '',
      };
      this.modalMarca = true;
    },
    cerrarModalMarca(){
      this.modalMarca = false;
      this.editandoMarcaId = null;
      this.nueva = {
        denominacion:'', clase:null, acta:'', estado:'',
        es_propia: true, titular:'',
        fecha_solicitud:'', fecha_publicacion:'', fecha_oposicion:'', fecha_concesion:'',
      };
    },
    async guardarMarca(){
      const url = this.editandoMarcaId
        ? '/api/dashboard/marcas/' + this.editandoMarcaId
        : '/api/dashboard/marcas';
      const method = this.editandoMarcaId ? 'PATCH' : 'POST';
      const r = await fetch(url, {
        method, headers:{'Content-Type':'application/json'},
        body: JSON.stringify(this.nueva),
      }).then(r=>r.json());
      if(!r.ok){ alert(r.error||'Error'); return; }
      this.cerrarModalMarca();
      await this.fetchMarcas();
    },

    async subirBulk(){
      if(!this.bulkFile) return;
      this.bulkLoading = true;
      this.bulkResult = null;
      try {
        const fd = new FormData();
        fd.append('file', this.bulkFile);
        const r = await fetch('/api/dashboard/marcas/bulk', {method:'POST', body: fd}).then(r=>r.json());
        if(!r.ok){ alert(r.error || 'Error'); return; }
        this.bulkResult = r.data;
        await this.fetchMarcas();
      } catch(e){
        alert('Error de red al subir el archivo.');
      } finally {
        this.bulkLoading = false;
      }
    },
    async iniciarVigilancia(marcaId){
      const r = await fetch('/api/dashboard/vigilancia/activar',{method:'POST',
        headers:{'Content-Type':'application/json'},
        body: JSON.stringify({marca_cliente_id: marcaId})}).then(r=>r.json());
      if(!r.ok){ alert(r.error||'Error'); return; }
      if(r.data.covered_by_premium){
        alert('Vigilancia activada (incluida en tu plan Premium).');
        await Promise.all([this.fetchVigilancia(), this.fetchMarcas()]);
        return;
      }
      if(r.data.init_point){ window.location.href = r.data.init_point; }
      else { await this.fetchVigilancia(); }
    },
    async cancelarVigilancia(id){
      if(!confirm('¿Cancelar la suscripción de vigilancia?')) return;
      const r = await fetch('/api/dashboard/vigilancia/'+id+'/cancelar',{method:'POST'}).then(r=>r.json());
      if(!r.ok){ alert(r.error||'Error'); return; }
      await this.fetchVigilancia();
    },
  };
}
</script>
</body></html>"""


@bp.route("/dashboard")
@premium_required
def dashboard_page():
    return render_template_string(DASHBOARD_PAGE)


# ─────────────────────────────────────────────────────────────────────
# API JSON del dashboard
# ─────────────────────────────────────────────────────────────────────

@bp.route("/api/dashboard/precios", methods=["GET"])
def api_precios():
    return _ok({
        "vigilancia_marca": PRECIO_VIGILANCIA_MARCA,
        "vigilancia_multi": PRECIO_VIGILANCIA_MULTI,
        "vigilancia_portfolio": PRECIO_VIGILANCIA_PORTFOLIO,
        "vigilancia_cap": PREMIUM_VIGILANCIA_CAP,
        "consulta_completa": float(os.getenv("PRECIO_CONSULTA_COMPLETA", "15000")),
    })


@bp.route("/api/dashboard/consultas", methods=["GET"])
@login_required
def api_consultas():
    user = current_user()
    with get_session() as s:
        rows = (s.query(Consulta)
                .filter((Consulta.user_id == user.id) | (Consulta.email == user.email))
                .order_by(Consulta.created_at.desc())
                .limit(100).all())
        return _ok([{
            "id": c.id, "marca": c.marca, "nivel": c.nivel,
            "diagnostico": c.diagnostico, "paid": c.paid,
            "created_at": c.created_at.isoformat(),
        } for c in rows])


@bp.route("/api/dashboard/marcas", methods=["GET"])
@login_required
def api_marcas_list():
    user = current_user()
    with get_session() as s:
        rows = s.query(MarcaCliente).filter_by(user_id=user.id).all()
        return _ok([_marca_payload(m) for m in rows])


def _marca_payload(m: "MarcaCliente") -> dict:
    """Serializa MarcaCliente y calcula DJU (5a) y fin del plazo de oposición (90d)."""
    fecha_dju = None
    if m.fecha_concesion:
        try:
            fecha_dju = m.fecha_concesion.replace(year=m.fecha_concesion.year + 5).isoformat()
        except ValueError:
            fecha_dju = m.fecha_concesion.replace(year=m.fecha_concesion.year + 5,
                                                  day=28).isoformat()

    fecha_oposicion_fin = None
    if m.fecha_publicacion:
        from datetime import timedelta as _td
        fecha_oposicion_fin = (m.fecha_publicacion + _td(days=90)).isoformat()

    return {
        "id": m.id, "denominacion": m.denominacion, "clase": m.clase,
        "acta": m.acta, "estado": m.estado, "titular": m.titular,
        "es_propia": bool(m.es_propia) if m.es_propia is not None else True,
        "fecha_solicitud": m.fecha_solicitud.isoformat() if m.fecha_solicitud else None,
        "fecha_publicacion": m.fecha_publicacion.isoformat() if m.fecha_publicacion else None,
        "fecha_oposicion": m.fecha_oposicion.isoformat() if m.fecha_oposicion else None,
        "fecha_oposicion_fin": fecha_oposicion_fin,
        "fecha_concesion": m.fecha_concesion.isoformat() if m.fecha_concesion else None,
        "fecha_vencimiento": m.fecha_vencimiento.isoformat() if m.fecha_vencimiento else None,
        "fecha_dju": fecha_dju,
    }


@bp.route("/api/dashboard/marcas", methods=["POST"])
@login_required
def api_marcas_add():
    user = current_user()
    data = request.get_json(silent=True) or {}
    deno = (data.get("denominacion") or "").strip()
    if not deno:
        return _err("Denominación requerida")

    def _parse_date(v):
        if not v:
            return None
        try:
            return datetime.fromisoformat(v).date()
        except Exception:
            return None

    fecha_solic = _parse_date(data.get("fecha_solicitud"))
    fecha_pub = _parse_date(data.get("fecha_publicacion"))
    fecha_opo = _parse_date(data.get("fecha_oposicion"))
    fecha_conc = _parse_date(data.get("fecha_concesion"))
    fecha_venc = _parse_date(data.get("fecha_vencimiento"))

    if fecha_conc and not fecha_venc:
        try:
            fecha_venc = fecha_conc.replace(year=fecha_conc.year + 10)
        except ValueError:
            fecha_venc = fecha_conc.replace(year=fecha_conc.year + 10, day=28)

    es_propia = data.get("es_propia")
    if es_propia is None:
        es_propia = True
    else:
        es_propia = bool(es_propia)

    with get_session() as s:
        m = MarcaCliente(
            user_id=user.id,
            denominacion=deno,
            clase=data.get("clase"),
            acta=(data.get("acta") or "").strip() or None,
            titular=(data.get("titular") or "").strip() or None,
            estado=(data.get("estado") or "").strip() or None,
            es_propia=es_propia,
            fecha_solicitud=fecha_solic,
            fecha_publicacion=fecha_pub,
            fecha_oposicion=fecha_opo,
            fecha_concesion=fecha_conc,
            fecha_vencimiento=fecha_venc,
            notas=(data.get("notas") or "").strip() or None,
        )
        s.add(m)
        s.commit()
        return _ok({"id": m.id}, 201)


@bp.route("/api/dashboard/marcas/template", methods=["GET"])
@login_required
def api_marcas_template():
    """Devuelve un XLSX template con logo, encabezados y filas de ejemplo."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import send_file

    wb = Workbook()
    ws = wb.active
    ws.title = "Mis marcas"

    columns = [
        ("denominacion", "denominación", "Nombre exacto de la marca", 28),
        ("clase", "clase", "Clase Niza (1-45)", 8),
        ("acta", "acta", "Nº de acta INPI (opcional)", 14),
        ("titular", "titular", "Solo si es marca de tercero", 22),
        ("estado", "estado", "solicitada / publicada / oposicion / concedida / vencida", 16),
        ("es_propia", "es_propia", "sí / no", 10),
        ("fecha_solicitud", "fecha_solicitud", "AAAA-MM-DD", 14),
        ("fecha_publicacion", "fecha_publicacion", "AAAA-MM-DD", 14),
        ("fecha_oposicion", "fecha_oposicion", "AAAA-MM-DD", 14),
        ("fecha_concesion", "fecha_concesion", "AAAA-MM-DD", 14),
        ("fecha_vencimiento", "fecha_vencimiento", "AAAA-MM-DD (se calcula sola si dejás concesión)", 16),
        ("notas", "notas", "Texto libre", 30),
    ]

    NAVY = "0D1B4B"
    BLUE = "1B6EF3"
    LIGHT_BLUE = "DBEAFE"
    GRAY = "64748B"
    LIGHT_GRAY = "F4F5F9"

    # Logo (si existe el archivo)
    try:
        from openpyxl.drawing.image import Image as XLImage
        logo_path = os.path.join(os.path.dirname(__file__), "..", "static", "logo-color.png")
        if os.path.exists(logo_path):
            img = XLImage(logo_path)
            img.height = 60
            img.width = 180
            ws.add_image(img, "A1")
            ws.row_dimensions[1].height = 50
            ws.row_dimensions[2].height = 20
    except Exception as e:
        logger.warning(f"No pude embedir logo en template: {e}")

    # Fila 3: título grande
    ws.merge_cells("A3:L3")
    ws["A3"] = "TEMPLATE — Carga masiva de marcas"
    ws["A3"].font = Font(name="Arial", size=18, bold=True, color=NAVY)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 30

    # Fila 4: instrucciones
    ws.merge_cells("A4:L4")
    ws["A4"] = ("Reemplazá las filas de ejemplo (5-7) por tus marcas. La primera columna "
                "(denominación) es obligatoria; el resto opcional. Guardá como .xlsx y subilo "
                "desde el panel de LegalPacers → Mis marcas → Cargar desde Excel.")
    ws["A4"].font = Font(name="Arial", size=10, color=GRAY, italic=True)
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 38

    # Fila 5: cabeceras de columnas (key + descripción debajo)
    header_row = 6
    desc_row = 7
    header_fill = PatternFill("solid", fgColor=BLUE)
    desc_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    border = Border(left=Side(style="thin", color="CBD5E1"),
                    right=Side(style="thin", color="CBD5E1"),
                    top=Side(style="thin", color="CBD5E1"),
                    bottom=Side(style="thin", color="CBD5E1"))

    for i, (key, label, hint, width) in enumerate(columns, start=1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = width

        c = ws.cell(row=header_row, column=i, value=key)
        c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

        d = ws.cell(row=desc_row, column=i, value=hint)
        d.font = Font(name="Arial", size=9, italic=True, color=NAVY)
        d.fill = desc_fill
        d.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        d.border = border

    ws.row_dimensions[header_row].height = 22
    ws.row_dimensions[desc_row].height = 28

    # Filas de ejemplo
    ejemplos = [
        ["Mi Marca",       25, "1234567",  None,        "concedida",   "sí", "2024-03-10", "2024-08-15", None,         "2025-02-04", None,         "Mi marca de ropa"],
        ["Acme",            9, None,       "Acme S.A.", "publicada",   "no", "2025-11-03", "2026-01-10", None,         None,         None,         "Competidor a vigilar"],
        ["Café Verbum",    30, "7654321",  None,        "solicitada",  "sí", "2026-02-20", None,         None,         None,         None,         ""],
    ]
    light = PatternFill("solid", fgColor=LIGHT_GRAY)
    for ri, fila in enumerate(ejemplos, start=desc_row + 1):
        for ci, val in enumerate(fila, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Arial", size=10, color="334155")
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = border
            c.fill = light

    # Freeze panes después de las cabeceras
    ws.freeze_panes = "A8"

    # Marca de agua tipo nota al final
    nota_row = desc_row + len(ejemplos) + 2
    ws.merge_cells(start_row=nota_row, start_column=1, end_row=nota_row, end_column=12)
    nota = ws.cell(row=nota_row, column=1,
                   value="Este archivo es solo un TEMPLATE de LegalPacers. Borrá las filas de ejemplo antes de subirlo.")
    nota.font = Font(name="Arial", size=10, italic=True, color=GRAY)
    nota.alignment = Alignment(horizontal="left", vertical="center")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="legalpacers-template-marcas.xlsx",
    )


@bp.route("/api/dashboard/marcas/bulk", methods=["POST"])
@login_required
def api_marcas_bulk():
    """Carga masiva de marcas vía Excel (.xlsx) o CSV.

    Columnas esperadas (case-insensitive, en cualquier orden):
      denominacion (obligatoria), clase, acta, titular, estado, es_propia,
      fecha_solicitud, fecha_publicacion, fecha_oposicion, fecha_concesion,
      fecha_vencimiento, notas

    Devuelve {creadas, errores: [{fila, motivo}]}.
    """
    user = current_user()
    f = request.files.get("file")
    if not f:
        return _err("Subí un archivo .xlsx o .csv")

    filename = (f.filename or "").lower()
    rows = []
    try:
        if filename.endswith(".csv"):
            import csv, io
            text = f.stream.read().decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            rows = [dict(r) for r in reader]
        elif filename.endswith(".xlsx") or filename.endswith(".xlsm"):
            from openpyxl import load_workbook
            wb = load_workbook(f.stream, read_only=True, data_only=True)
            ws = wb.active
            headers = None
            for r in ws.iter_rows(values_only=True):
                if not any(r):
                    continue
                if headers is None:
                    headers = [str(c).strip().lower() if c is not None else f"col{i}"
                               for i, c in enumerate(r)]
                    continue
                row = {headers[i]: r[i] for i in range(min(len(headers), len(r)))}
                rows.append(row)
        else:
            return _err("Formato no soportado. Usá .xlsx o .csv")
    except Exception as e:
        logger.exception("Error parseando bulk upload")
        return _err(f"No pudimos leer el archivo: {e}")

    if not rows:
        return _err("El archivo está vacío")

    def _to_date(v):
        if v is None or v == "":
            return None
        if isinstance(v, datetime):
            return v.date()
        if hasattr(v, "year") and hasattr(v, "month"):  # date
            return v
        try:
            return datetime.fromisoformat(str(v).strip()[:10]).date()
        except Exception:
            for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%m/%d/%Y"):
                try:
                    return datetime.strptime(str(v).strip(), fmt).date()
                except Exception:
                    continue
            return None

    def _to_bool_propia(v):
        if v is None or v == "":
            return True
        s = str(v).strip().lower()
        if s in ("no", "tercero", "false", "0", "n"):
            return False
        return True

    creadas = 0
    errores = []
    with get_session() as s:
        for idx, row in enumerate(rows, start=2):  # fila 1 = headers
            r = {(k or "").strip().lower(): v for k, v in row.items()}
            deno = (r.get("denominacion") or r.get("denominación") or r.get("marca") or "").strip() if isinstance(r.get("denominacion") or r.get("denominación") or r.get("marca"), str) else r.get("denominacion") or r.get("denominación") or r.get("marca")
            if not deno or not str(deno).strip():
                errores.append({"fila": idx, "motivo": "Falta denominación"})
                continue
            try:
                clase_raw = r.get("clase")
                clase_int = int(clase_raw) if clase_raw not in (None, "") else None
            except (ValueError, TypeError):
                clase_int = None
            try:
                m = MarcaCliente(
                    user_id=user.id,
                    denominacion=str(deno).strip()[:300],
                    clase=clase_int,
                    acta=(str(r.get("acta") or "").strip() or None),
                    titular=(str(r.get("titular") or "").strip() or None),
                    estado=(str(r.get("estado") or "").strip() or None),
                    es_propia=_to_bool_propia(r.get("es_propia") or r.get("propia")),
                    fecha_solicitud=_to_date(r.get("fecha_solicitud") or r.get("inicio")),
                    fecha_publicacion=_to_date(r.get("fecha_publicacion") or r.get("publicacion") or r.get("publicación")),
                    fecha_oposicion=_to_date(r.get("fecha_oposicion") or r.get("oposicion") or r.get("oposición")),
                    fecha_concesion=_to_date(r.get("fecha_concesion") or r.get("concesion") or r.get("concesión")),
                    fecha_vencimiento=_to_date(r.get("fecha_vencimiento") or r.get("vencimiento")),
                    notas=(str(r.get("notas") or "").strip() or None),
                )
                s.add(m)
                creadas += 1
            except Exception as e:
                errores.append({"fila": idx, "motivo": str(e)[:200]})
        s.commit()

    return _ok({"creadas": creadas, "errores": errores})


@bp.route("/api/dashboard/marcas/<int:marca_id>", methods=["PATCH", "PUT"])
@login_required
def api_marcas_update(marca_id: int):
    """Edita los campos de una marca existente. Acepta cualquier subset de campos."""
    user = current_user()
    data = request.get_json(silent=True) or {}

    def _parse_date(v):
        if v in (None, ""):
            return None
        if hasattr(v, "year") and hasattr(v, "month"):  # ya es date
            return v
        try:
            return datetime.fromisoformat(str(v)[:10]).date()
        except Exception:
            return None

    with get_session() as s:
        m = s.query(MarcaCliente).filter_by(id=marca_id, user_id=user.id).first()
        if not m:
            return _err("No encontrada", 404)

        # String fields: actualizamos solo si vinieron en el payload
        for field in ("denominacion", "acta", "titular", "estado", "notas"):
            if field in data:
                v = (data.get(field) or "").strip() if isinstance(data.get(field), str) else data.get(field)
                setattr(m, field, v or None)

        if "clase" in data:
            try:
                m.clase = int(data["clase"]) if data["clase"] not in (None, "") else None
            except (ValueError, TypeError):
                pass

        if "es_propia" in data:
            m.es_propia = bool(data["es_propia"])

        for date_field in ("fecha_solicitud", "fecha_publicacion", "fecha_oposicion",
                            "fecha_concesion", "fecha_vencimiento"):
            if date_field in data:
                setattr(m, date_field, _parse_date(data[date_field]))

        # Si actualizan fecha_concesion y NO mandan fecha_vencimiento, recalculamos
        if "fecha_concesion" in data and "fecha_vencimiento" not in data and m.fecha_concesion:
            try:
                m.fecha_vencimiento = m.fecha_concesion.replace(year=m.fecha_concesion.year + 10)
            except ValueError:
                m.fecha_vencimiento = m.fecha_concesion.replace(year=m.fecha_concesion.year + 10, day=28)

        s.commit()
        s.refresh(m)
        return _ok(_marca_payload(m))


@bp.route("/api/dashboard/marcas/<int:marca_id>", methods=["DELETE"])
@login_required
def api_marcas_delete(marca_id: int):
    user = current_user()
    with get_session() as s:
        m = s.query(MarcaCliente).filter_by(id=marca_id, user_id=user.id).first()
        if not m:
            return _err("No encontrada", 404)
        s.delete(m)
        s.commit()
        return _ok({"deleted": True})


@bp.route("/api/dashboard/vigilancia", methods=["GET"])
@login_required
def api_vigilancia_list():
    user = current_user()
    with get_session() as s:
        rows = (s.query(SuscripcionVigilancia)
                .filter_by(user_id=user.id).order_by(SuscripcionVigilancia.activated_at.desc()).all())
        out = []
        for v in rows:
            marca_nombre = None
            if v.marca_cliente_id:
                mc = s.query(MarcaCliente).filter_by(id=v.marca_cliente_id).first()
                marca_nombre = mc.denominacion if mc else None
            out.append({
                "id": v.id,
                "marca_cliente_id": v.marca_cliente_id,
                "marca_nombre": marca_nombre,
                "tipo": v.tipo,
                "status": v.status,
                "monto": v.monto,
                "activated_at": v.activated_at.isoformat() if v.activated_at else None,
                "next_check_at": v.next_check_at.isoformat() if v.next_check_at else None,
            })
        return _ok(out)


PREMIUM_VIGILANCIA_CAP = int(os.getenv("PREMIUM_VIGILANCIA_CAP", "10"))


@bp.route("/api/dashboard/vigilancia/activar", methods=["POST"])
@login_required
def api_vigilancia_activar():
    user = current_user()
    data = request.get_json(silent=True) or {}
    marca_cliente_id = data.get("marca_cliente_id")
    tipo = data.get("tipo", "marca")  # 'marca' o 'portfolio'

    with get_session() as s:
        if tipo == "marca":
            if not marca_cliente_id:
                return _err("Seleccioná una marca para vigilar")
            mc = s.query(MarcaCliente).filter_by(
                id=marca_cliente_id, user_id=user.id).first()
            if not mc:
                return _err("Marca no encontrada", 404)
            descripcion = f"Vigilancia mensual de \"{mc.denominacion}\""
            monto = PRECIO_VIGILANCIA_MARCA
        else:
            descripcion = "Vigilancia mensual de portfolio (todas mis marcas)"
            monto = PRECIO_VIGILANCIA_PORTFOLIO

        # Evitar duplicados activos
        existing = (s.query(SuscripcionVigilancia)
                    .filter_by(user_id=user.id, marca_cliente_id=marca_cliente_id,
                               status="active").first())
        if existing:
            return _err("Ya tenés vigilancia activa sobre esa marca", 409)

        # ¿Admin? Acceso full sin caps ni cobros.
        if user.is_admin:
            sub = SuscripcionVigilancia(
                user_id=user.id,
                marca_cliente_id=marca_cliente_id,
                tipo=tipo, status="active", monto=0.0,
                activated_at=datetime.utcnow(),
                metadata_json={"covered_by": "admin"},
            )
            s.add(sub)
            s.commit()
            s.refresh(sub)
            return _ok({
                "suscripcion_id": sub.id,
                "monto": 0.0,
                "covered_by_premium": True,
                "init_point": None,
            })

        # ¿Premium activo? — entonces vigilancia gratis hasta el cap (10 marcas)
        premium = has_active_premium(user)
        vigiladas_count = (s.query(SuscripcionVigilancia)
                           .filter_by(user_id=user.id, status="active", tipo="marca")
                           .count())
        covered_by_premium = premium and vigiladas_count < PREMIUM_VIGILANCIA_CAP

        if covered_by_premium:
            # Activación inmediata, sin pasar por MP
            sub = SuscripcionVigilancia(
                user_id=user.id,
                marca_cliente_id=marca_cliente_id,
                tipo=tipo, status="active", monto=0.0,
                activated_at=datetime.utcnow(),
                metadata_json={"covered_by": "premium"},
            )
            s.add(sub)
            s.commit()
            s.refresh(sub)
            return _ok({
                "suscripcion_id": sub.id,
                "monto": 0.0,
                "covered_by_premium": True,
                "init_point": None,
            })

        # Premium con cap superado → vigilancia adicional al precio extra.
        # No bloqueamos: cobramos $1.500/mes (configurable) y sigue.

        sub = SuscripcionVigilancia(
            user_id=user.id,
            marca_cliente_id=marca_cliente_id,
            tipo=tipo, status="pending", monto=monto,
        )
        s.add(sub)
        s.commit()
        s.refresh(sub)
        sub_id = sub.id

    from services.mercadopago import create_vigilancia_subscription
    pref = create_vigilancia_subscription(
        suscripcion_id=sub_id, email=user.email,
        monto=monto, descripcion=descripcion,
    )

    return _ok({
        "suscripcion_id": sub_id,
        "monto": monto,
        "covered_by_premium": False,
        "init_point": pref.get("init_point"),
        "preapproval_id": pref.get("id"),
    })


@bp.route("/api/dashboard/vigilancia/<int:sub_id>/cancelar", methods=["POST"])
@login_required
def api_vigilancia_cancelar(sub_id: int):
    user = current_user()
    paused_covered = 0
    with get_session() as s:
        sub = (s.query(SuscripcionVigilancia)
               .filter_by(id=sub_id, user_id=user.id).first())
        if not sub:
            return _err("No encontrada", 404)
        if sub.status == "cancelled":
            return _ok({"already_cancelled": True})

        from services.mercadopago import cancel_subscription
        if sub.mp_subscription_id:
            cancel_subscription(sub.mp_subscription_id)

        sub.status = "cancelled"
        sub.cancelled_at = datetime.utcnow()

        # Si cancela el Premium, pausamos todas las vigilancias cubiertas por el plan.
        if sub.tipo == "premium":
            covered = (s.query(SuscripcionVigilancia)
                       .filter_by(user_id=user.id, status="active")
                       .filter(SuscripcionVigilancia.id != sub.id).all())
            for v in covered:
                md = v.metadata_json or {}
                if md.get("covered_by") == "premium":
                    v.status = "paused"
                    v.paused_at = datetime.utcnow()
                    paused_covered += 1

        s.commit()
        return _ok({"cancelled": True, "paused_covered": paused_covered})


@bp.route("/api/dashboard/alertas", methods=["GET"])
@login_required
def api_alertas():
    user = current_user()
    with get_session() as s:
        rows = (s.query(AlertaVigilancia)
                .filter_by(user_id=user.id)
                .order_by(AlertaVigilancia.created_at.desc())
                .limit(200).all())
        out = []
        for a in rows:
            mc = (s.query(MarcaCliente).filter_by(id=a.marca_cliente_id).first()
                  if a.marca_cliente_id else None)
            out.append({
                "id": a.id, "score": a.score, "nivel": a.nivel,
                "marca_propia": mc.denominacion if mc else "—",
                "marca_nueva_denominacion": a.marca_nueva_denominacion,
                "marca_nueva_clase": a.marca_nueva_clase,
                "marca_nueva_titular": a.marca_nueva_titular,
                "marca_nueva_acta": a.marca_nueva_acta,
                "boletin_num": a.boletin_num,
                "created_at": a.created_at.isoformat(),
            })
        return _ok(out)


@bp.route("/api/dashboard/premium", methods=["GET"])
@login_required
def api_premium_info():
    """Devuelve la suscripción premium activa del usuario, si tiene."""
    user = current_user()
    with get_session() as s:
        sub = (s.query(SuscripcionVigilancia)
               .filter_by(user_id=user.id, tipo="premium")
               .filter(SuscripcionVigilancia.status.in_(["active", "paused"]))
               .order_by(SuscripcionVigilancia.activated_at.desc())
               .first())
        if not sub:
            return _ok(None)
        return _ok({
            "id": sub.id,
            "status": sub.status,
            "plan_freq": sub.plan_freq or "mensual",
            "monto": sub.monto,
            "auto_renew": bool(sub.auto_renew),
            "paid_through_date": sub.paid_through_date.isoformat() if sub.paid_through_date else None,
            "activated_at": sub.activated_at.isoformat() if sub.activated_at else None,
        })


@bp.route("/api/dashboard/premium/auto_renew", methods=["POST"])
@login_required
def api_premium_auto_renew():
    """Activa o desactiva la renovación automática del Premium del usuario."""
    user = current_user()
    data = request.get_json(silent=True) or {}
    new_value = bool(data.get("auto_renew", True))

    with get_session() as s:
        sub = (s.query(SuscripcionVigilancia)
               .filter_by(user_id=user.id, tipo="premium", status="active")
               .order_by(SuscripcionVigilancia.activated_at.desc())
               .first())
        if not sub:
            return _err("No tenés una suscripción premium activa", 404)
        sub.auto_renew = new_value
        s.commit()

        # Si desactivó la renovación, cancelamos el preapproval en MP para
        # que no haya más cobros. La suscripción local sigue 'active' hasta
        # paid_through (lo maneja run_subscription_maintenance).
        if not new_value and sub.mp_subscription_id:
            try:
                from services.mercadopago import cancel_subscription
                cancel_subscription(sub.mp_subscription_id)
            except Exception as e:
                logger.warning(f"No pude cancelar MP preapproval {sub.mp_subscription_id}: {e}")

    return _ok({"auto_renew": new_value})


PROFILE_FIELDS_STR = [
    "dni", "direccion", "localidad", "provincia", "pais",
    "instagram", "linkedin", "twitter", "facebook", "tiktok", "web",
    "empresa_nombre", "empresa_rol", "empresa_industria", "empresa_tamano", "empresa_cuit",
    "bio", "ofrece", "necesidades", "avatar_data_uri",
]


@bp.route("/api/dashboard/perfil", methods=["GET"])
@login_required
def api_perfil_get():
    user = current_user()
    with get_session() as s:
        p = s.query(UserProfile).filter_by(user_id=user.id).first()
        if not p:
            return _ok(None)
        out = {f: getattr(p, f) or "" for f in PROFILE_FIELDS_STR}
        out["fecha_nacimiento"] = p.fecha_nacimiento.isoformat() if p.fecha_nacimiento else ""
        return _ok(out)


@bp.route("/api/dashboard/perfil", methods=["POST"])
@login_required
def api_perfil_save():
    user = current_user()
    data = request.get_json(silent=True) or {}

    def _trim(v):
        return (v or "").strip() if isinstance(v, str) else v

    # Validar tamaño del avatar (~ <300KB en base64)
    avatar = (data.get("avatar_data_uri") or "").strip()
    if avatar and len(avatar) > 500_000:
        return _err("La foto es demasiado grande. Probá una más chica.")

    with get_session() as s:
        p = s.query(UserProfile).filter_by(user_id=user.id).first()
        if not p:
            p = UserProfile(user_id=user.id)
            s.add(p)

        for f in PROFILE_FIELDS_STR:
            if f in data:
                setattr(p, f, _trim(data[f]) or None)

        if "fecha_nacimiento" in data:
            v = data.get("fecha_nacimiento")
            try:
                p.fecha_nacimiento = datetime.fromisoformat(v).date() if v else None
            except Exception:
                p.fecha_nacimiento = None

        s.commit()
    return _ok({"saved": True})


@bp.route("/api/dashboard/config", methods=["POST"])
@login_required
def api_config():
    user = current_user()
    data = request.get_json(silent=True) or {}
    nombre = (data.get("nombre") or "").strip() or None
    telefono = (data.get("telefono") or "").strip() or None
    alertas_wa = bool(data.get("alertas_whatsapp"))

    if alertas_wa and not telefono:
        return _err("Cargá un teléfono para activar alertas por WhatsApp")

    with get_session() as s:
        from database import User
        u = s.query(User).filter_by(id=user.id).first()
        if not u:
            return _err("Usuario no encontrado", 404)
        u.nombre = nombre
        u.telefono = telefono
        u.alertas_whatsapp = alertas_wa
        s.commit()
    return _ok({"saved": True})


@bp.route("/api/dashboard/pagos", methods=["GET"])
@login_required
def api_pagos():
    user = current_user()
    with get_session() as s:
        rows = (s.query(Pago)
                .filter((Pago.user_id == user.id) | (Pago.email == user.email))
                .order_by(Pago.created_at.desc())
                .limit(100).all())
        return _ok([{
            "id": p.id, "tipo": p.tipo, "monto": p.monto, "moneda": p.moneda,
            "status": p.status,
            "created_at": p.created_at.isoformat(),
            "paid_at": p.paid_at.isoformat() if p.paid_at else None,
        } for p in rows])
